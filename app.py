#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Kengni Finance - Complete Financial Management & Trading Application
Version 2.0 - Enhanced with AI Analysis and Advanced Features
Migration: SQLite → MongoDB (Vercel deployment)
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, flash
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import os
from datetime import datetime, timedelta
import secrets
import random
import json
import pandas as pd
from io import BytesIO
import yfinance as yf
import numpy as np
import base64
from PIL import Image
import io
import urllib.request
import re
import smtplib
import threading
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# ── Import couche MongoDB ──────────────────────────────────────────────────────
from mongo_db import (
    get_mongo_db, get_col, get_next_id, doc_to_dict, docs_to_list, init_db
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'static/uploads'

# ── Configuration Gmail pour les rappels d'agenda ──────────────────────────────
GMAIL_CONFIG = {
    'sender_email':    os.environ.get('GMAIL_SENDER',       'fabrice.kengni12@gmail.com'),
    'sender_name':     'Kengni Finance — Agenda',
    'receiver_email':  os.environ.get('GMAIL_RECEIVER',     'fabrice.kengni@icloud.com'),
    'smtp_host':       'smtp.gmail.com',
    'smtp_port':       587,
    'smtp_password':   os.environ.get('GMAIL_APP_PASSWORD', ''),
}

# ── Types et couleurs des événements d'agenda ─────────────────────────────────
AGENDA_EVENT_COLORS = {
    'trading':   {'bg': '#00d4aa', 'border': '#00b894', 'icon': '📈', 'label': 'Trading'},
    'finance':   {'bg': '#4a9eff', 'border': '#2980b9', 'icon': '💰', 'label': 'Finance'},
    'formation': {'bg': '#a29bfe', 'border': '#6c5ce7', 'icon': '📚', 'label': 'Formation'},
    'personnel': {'bg': '#fd79a8', 'border': '#e84393', 'icon': '👤', 'label': 'Personnel'},
    'reunion':   {'bg': '#ffd700', 'border': '#f39c12', 'icon': '🤝', 'label': 'Réunion'},
    'revue':     {'bg': '#ff7675', 'border': '#d63031', 'icon': '🔍', 'label': 'Revue'},
}

# ── Informations de paiement Kengni Trading Academy ─────────────────────────
PAYMENT_INFO = {
    'orange_money': {'numero': '695 072 659', 'nom': 'Fabrice Kengni', 'label': 'Orange Money'},
    'mtn_money':    {'numero': '670 695 946', 'nom': 'Fabrice Kengni', 'label': 'MTN MoMo'},
    'paypal':       {'adresse': 'fabrice.kengni@icloud.com', 'label': 'PayPal'},
    'crypto':       {'adresse': 'fabrice.kengni@icloud.com', 'label': 'Crypto (via email)'},
}

FORMATION_PRICES = {
    'Débutant':       {'xaf': 25000,  'eur': 38},
    'Intermédiaire':  {'xaf': 50000,  'eur': 76},
    'Avancé':         {'xaf': 100000, 'eur': 152},
    'Pro / Mentoring':{'xaf': 200000, 'eur': 305},
}

# Add Python built-in functions to Jinja2 environment
app.jinja_env.globals.update({
    'abs': abs,
    'min': min,
    'max': max,
    'round': round,
    'int': int,
    'float': float,
    'len': len,
    'sum': sum
})

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Allowed extensions for image uploads
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ── URL secrète admin ──
ADMIN_SECRET_TOKEN        = os.environ.get('ADMIN_SECRET_TOKEN',        'kengni-control-7749')
ADMIN_SECONDARY_PASSWORD  = os.environ.get('ADMIN_SECONDARY_PASSWORD',  'Kengni@fablo12')

# ══════════════════════════════════════════════════════════════
# HELPERS DB
# ══════════════════════════════════════════════════════════════

def db():
    return get_mongo_db()

def _date_ago(days=30):
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

def _now_iso():
    return datetime.now().isoformat()

# ══════════════════════════════════════════════════════════════
# CONTEXT PROCESSOR
# ══════════════════════════════════════════════════════════════

@app.context_processor
def inject_global_context():
    ctx = {'training_total_nav': 0}
    if 'user_id' in session:
        try:
            user_id = session['user_id']
            count = get_col('training_courses').count_documents({"user_id": user_id})
            ctx['training_total_nav'] = count
        except Exception:
            pass
    return ctx

# ══════════════════════════════════════════════════════════════
# DECORATORS
# ══════════════════════════════════════════════════════════════

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') not in ('admin', 'superadmin'):
            from flask import abort; abort(404)
        return f(*args, **kwargs)
    return decorated_function

# ══════════════════════════════════════════════════════════════
# NOTIFICATION HELPER
# ══════════════════════════════════════════════════════════════

def create_notification(user_id, notif_type, title, message):
    try:
        nid = get_next_id("notifications")
        get_col('notifications').insert_one({
            "id": nid,
            "user_id": user_id,
            "type": notif_type,
            "title": title,
            "message": message,
            "is_read": 0,
            "action_url": None,
            "created_at": _now_iso()
        })
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════
# AI ANALYSIS FUNCTIONS
# ══════════════════════════════════════════════════════════════

def analyze_trading_psychology(user_id):
    patterns = []
    try:
        transactions = docs_to_list(
            get_col('transactions').find(
                {"user_id": user_id},
                sort=[("created_at", -1)],
                limit=50
            )
        )
        journal_entries = docs_to_list(
            get_col('trading_journal').find(
                {"user_id": user_id},
                sort=[("created_at", -1)],
                limit=20
            )
        )

        # 1. Overtrading detection
        recent_24h = sum(1 for t in transactions
                         if datetime.fromisoformat(t['created_at']) > datetime.now() - timedelta(hours=24))
        if recent_24h > 10:
            patterns.append({
                'type': 'overtrading',
                'severity': 'high' if recent_24h > 20 else 'medium',
                'description': f'Vous avez effectué {recent_24h} transactions en 24h',
                'recommendation': 'Prenez du recul. Le overtrading augmente les frais et diminue la qualité des décisions.'
            })

        # 2. FOMO detection
        buy_after_loss = 0
        for i in range(1, min(len(transactions), 10)):
            if transactions[i]['type'] == 'buy' and i > 0:
                prev = transactions[i-1]
                if prev['type'] == 'sell' and prev['amount'] < 0:
                    buy_after_loss += 1
        if buy_after_loss >= 3:
            patterns.append({
                'type': 'FOMO',
                'severity': 'high',
                'description': 'Tendance à acheter immédiatement après des pertes',
                'recommendation': 'Attendez 30 minutes avant toute nouvelle transaction après une perte.'
            })

        # 3. Revenge trading
        consecutive_losses = 0
        max_consecutive = 0
        for t in transactions:
            if t['type'] == 'sell' and t['amount'] < 0:
                consecutive_losses += 1
                max_consecutive = max(max_consecutive, consecutive_losses)
            else:
                consecutive_losses = 0
        if max_consecutive >= 3:
            patterns.append({
                'type': 'revenge_trading',
                'severity': 'critical',
                'description': f'{max_consecutive} pertes consécutives détectées',
                'recommendation': 'Arrêtez de trader après 2 pertes consécutives. Analysez vos erreurs.'
            })

        # 4. Emotional patterns from journal
        emotional_keywords = {
            'fear': ['peur', 'anxieux', 'stressé', 'inquiet', 'nerveux'],
            'greed': ['avidité', 'cupide', 'trop confiant', 'sûr de moi'],
            'overconfidence': ['facile', 'certain', 'évident', 'garanti']
        }
        for entry in journal_entries:
            if entry.get('emotions'):
                emotions_text = entry['emotions'].lower()
                for emotion, keywords in emotional_keywords.items():
                    if any(kw in emotions_text for kw in keywords):
                        patterns.append({
                            'type': emotion,
                            'severity': 'medium',
                            'description': f'Émotion détectée: {emotion}',
                            'recommendation': 'Identifiée dans votre journal. Restez objectif.'
                        })

        # Save patterns to MongoDB
        for pattern in patterns:
            pid = get_next_id("psychological_patterns")
            get_col('psychological_patterns').insert_one({
                "id": pid,
                "user_id": user_id,
                "pattern_type": pattern['type'],
                "severity": pattern['severity'],
                "detected_date": _now_iso(),
                "description": pattern['description'],
                "recommendations": pattern['recommendation'],
                "status": "active"
            })
    except Exception as e:
        print(f"analyze_trading_psychology error: {e}")
    return patterns


def calculate_trader_score(user_id):
    score_data = {
        'overall_score': 50,
        'discipline_score': 50,
        'risk_management_score': 50,
        'strategy_consistency_score': 50,
        'emotional_control_score': 50,
        'profitability_score': 50,
        'details': {}
    }
    try:
        transactions = docs_to_list(
            get_col('transactions').find(
                {"user_id": user_id},
                sort=[("created_at", -1)],
                limit=100
            )
        )
        if not transactions:
            return score_data

        # 1. Profitability Score
        total_profit = sum(t['amount'] for t in transactions if t['type'] == 'sell')
        total_invested = sum(abs(t['amount']) for t in transactions if t['type'] == 'buy')
        if total_invested > 0:
            roi = (total_profit / total_invested) * 100
            score_data['profitability_score'] = min(100, max(0, 50 + roi))
        wins = sum(1 for t in transactions if t['type'] == 'sell' and t['amount'] > 0)
        total_sells = sum(1 for t in transactions if t['type'] == 'sell')
        win_rate = (wins / total_sells * 100) if total_sells > 0 else 0
        score_data['details']['win_rate'] = round(win_rate, 2)

        # 2. Risk Management Score
        risk_score = 50
        positions_with_sl = get_col('positions').count_documents(
            {"user_id": user_id, "stop_loss": {"$ne": None}}
        )
        total_positions = get_col('positions').count_documents({"user_id": user_id})
        if total_positions > 0:
            sl_percentage = (positions_with_sl / total_positions) * 100
            risk_score += (sl_percentage - 50) * 0.5
        amounts = [abs(t['amount']) for t in transactions if t['type'] == 'buy']
        if len(amounts) > 5:
            avg_amount = np.mean(amounts)
            std_amount = np.std(amounts)
            cv = (std_amount / avg_amount) if avg_amount > 0 else 0
            if cv < 0.3:
                risk_score += 20
            elif cv > 0.8:
                risk_score -= 20
        score_data['risk_management_score'] = min(100, max(0, risk_score))

        # 3. Discipline Score
        discipline_score = 50
        recent_24h = sum(1 for t in transactions
                         if datetime.fromisoformat(t['created_at']) > datetime.now() - timedelta(hours=24))
        if recent_24h > 15:
            discipline_score -= 30
        elif recent_24h < 5:
            discipline_score += 20
        revenge_patterns = get_col('psychological_patterns').count_documents(
            {"user_id": user_id, "pattern_type": "revenge_trading", "status": "active"}
        )
        if revenge_patterns > 0:
            discipline_score -= 20
        score_data['discipline_score'] = min(100, max(0, discipline_score))

        # 4. Strategy Consistency
        pipeline = [
            {"$match": {"user_id": user_id, "strategy": {"$ne": None}}},
            {"$group": {"_id": "$strategy", "count": {"$sum": 1}}}
        ]
        strategies = list(get_col('transactions').aggregate(pipeline))
        strategy_score = 50
        if strategies:
            max_strategy_count = max(s['count'] for s in strategies)
            total_with_strategy = sum(s['count'] for s in strategies)
            consistency = (max_strategy_count / total_with_strategy * 100) if total_with_strategy > 0 else 0
            strategy_score = min(100, consistency)
        score_data['strategy_consistency_score'] = strategy_score

        # 5. Emotional Control
        active_patterns = get_col('psychological_patterns').count_documents(
            {"user_id": user_id, "status": "active"}
        )
        emotional_score = 100 - (active_patterns * 15)
        score_data['emotional_control_score'] = min(100, max(0, emotional_score))

        # Overall score
        score_data['overall_score'] = round(
            score_data['profitability_score'] * 0.30 +
            score_data['risk_management_score'] * 0.25 +
            score_data['discipline_score'] * 0.20 +
            score_data['strategy_consistency_score'] * 0.15 +
            score_data['emotional_control_score'] * 0.10,
            2
        )

        # Save score
        sid = get_next_id("trader_scores")
        get_col('trader_scores').insert_one({
            "id": sid,
            "user_id": user_id,
            "date": _now_iso(),
            "overall_score": score_data['overall_score'],
            "discipline_score": score_data['discipline_score'],
            "risk_management_score": score_data['risk_management_score'],
            "strategy_consistency_score": score_data['strategy_consistency_score'],
            "emotional_control_score": score_data['emotional_control_score'],
            "profitability_score": score_data['profitability_score'],
            "monthly_trades": len(transactions),
            "win_rate": win_rate,
            "created_at": _now_iso()
        })
    except Exception as e:
        print(f"calculate_trader_score error: {e}")
    return score_data


def analyze_financial_report(data):
    try:
        insights = {
            'summary': "Analyse automatique du rapport financier",
            'recommendations': [],
            'risks': [],
            'opportunities': [],
            'anomalies': []
        }
        if 'revenue' in data and 'expenses' in data:
            profit_margin = ((data['revenue'] - data['expenses']) / data['revenue'] * 100) if data['revenue'] > 0 else 0
            if profit_margin > 20:
                insights['recommendations'].append("✅ Excellente marge bénéficiaire. Envisagez d'investir dans la croissance.")
                insights['opportunities'].append("Capacité d'investissement disponible")
            elif profit_margin < 10:
                insights['recommendations'].append("⚠️ Marge bénéficiaire faible. Optimisez vos dépenses.")
                insights['risks'].append("Risque de rentabilité")
            if data['expenses'] > data['revenue']:
                insights['risks'].append("🚨 Dépenses supérieures aux revenus - attention critique!")
                insights['recommendations'].append("Action immédiate requise: réduire les dépenses de " +
                    f"{round((data['expenses'] - data['revenue']) / data['revenue'] * 100, 2)}%")
            if data['expenses'] > data['revenue'] * 1.5:
                insights['anomalies'].append("Dépenses anormalement élevées détectées")
        return insights
    except Exception as e:
        return {'error': str(e)}


def analyze_trade_image(image_path, trade_data):
    insights = {
        'setup_quality': 'N/A',
        'entry_timing': 'N/A',
        'risk_reward': 'N/A',
        'recommendations': []
    }
    try:
        if trade_data.get('risk_reward_ratio'):
            rr = trade_data['risk_reward_ratio']
            if rr >= 2:
                insights['risk_reward'] = 'Excellent'
                insights['recommendations'].append('✅ Bon ratio risque/récompense')
            elif rr >= 1:
                insights['risk_reward'] = 'Acceptable'
                insights['recommendations'].append('⚠️ Ratio risque/récompense minimum atteint')
            else:
                insights['risk_reward'] = 'Mauvais'
                insights['recommendations'].append('❌ Ratio risque/récompense insuffisant')
        if trade_data.get('profit_loss'):
            pl = trade_data['profit_loss']
            if pl > 0:
                insights['recommendations'].append('✅ Trade gagnant - analysez ce qui a fonctionné')
            else:
                insights['recommendations'].append('📝 Trade perdant - identifiez les erreurs')
        if trade_data.get('strategy'):
            insights['setup_quality'] = 'Défini'
            insights['recommendations'].append(f'Strategy utilisée: {trade_data["strategy"]}')
    except Exception as e:
        insights['error'] = str(e)
    return insights


def trading_recommendation(symbol, timeframe='1mo'):
    try:
        ticker = yf.Ticker(symbol)
        hist = ticker.history(period=timeframe)
        if hist.empty:
            return {'error': 'Données non disponibles'}
        current_price = hist['Close'].iloc[-1]
        sma_20 = hist['Close'].rolling(window=20).mean().iloc[-1]
        sma_50 = hist['Close'].rolling(window=50).mean().iloc[-1] if len(hist) >= 50 else sma_20
        rsi = calculate_rsi(hist['Close'])
        avg_volume = hist['Volume'].mean()
        recent_volume = hist['Volume'].iloc[-1]
        volume_ratio = recent_volume / avg_volume if avg_volume > 0 else 1
        recommendation = {
            'symbol': symbol,
            'current_price': round(current_price, 2),
            'sma_20': round(sma_20, 2),
            'sma_50': round(sma_50, 2),
            'rsi': round(rsi, 2),
            'volume_ratio': round(volume_ratio, 2),
            'signal': 'NEUTRE',
            'confidence': 50,
            'analysis': [],
            'entry_points': [],
            'stop_loss': 0,
            'take_profit': 0
        }
        if current_price > sma_20 and sma_20 > sma_50:
            recommendation['analysis'].append("📈 Tendance haussière confirmée")
            trend_score = 20
        elif current_price < sma_20 and sma_20 < sma_50:
            recommendation['analysis'].append("📉 Tendance baissière confirmée")
            trend_score = -20
        else:
            recommendation['analysis'].append("➡️ Tendance neutre/consolidation")
            trend_score = 0
        if rsi > 70:
            recommendation['analysis'].append(f"🔴 RSI: {round(rsi, 2)} - Surachat détecté")
            rsi_score = -15
        elif rsi < 30:
            recommendation['analysis'].append(f"🟢 RSI: {round(rsi, 2)} - Survente détectée")
            rsi_score = 15
        else:
            recommendation['analysis'].append(f"🟡 RSI: {round(rsi, 2)} - Zone neutre")
            rsi_score = 0
        if volume_ratio > 1.5:
            recommendation['analysis'].append(f"📊 Volume élevé ({round(volume_ratio, 2)}x) - Signal fort")
            volume_score = 10
        else:
            volume_score = 0
        total_score = trend_score + rsi_score + volume_score
        if total_score > 20:
            recommendation['signal'] = 'ACHAT FORT'
            recommendation['confidence'] = min(90, 50 + total_score)
            recommendation['entry_points'].append(round(current_price * 0.98, 2))
            recommendation['stop_loss'] = round(current_price * 0.95, 2)
            recommendation['take_profit'] = round(current_price * 1.10, 2)
        elif total_score > 5:
            recommendation['signal'] = 'ACHAT'
            recommendation['confidence'] = min(80, 50 + total_score)
            recommendation['entry_points'].append(round(current_price * 0.99, 2))
            recommendation['stop_loss'] = round(current_price * 0.96, 2)
            recommendation['take_profit'] = round(current_price * 1.08, 2)
        elif total_score < -20:
            recommendation['signal'] = 'VENTE FORTE'
            recommendation['confidence'] = min(90, 50 + abs(total_score))
            recommendation['entry_points'].append(round(current_price * 1.02, 2))
            recommendation['stop_loss'] = round(current_price * 1.05, 2)
            recommendation['take_profit'] = round(current_price * 0.90, 2)
        elif total_score < -5:
            recommendation['signal'] = 'VENTE'
            recommendation['confidence'] = min(80, 50 + abs(total_score))
            recommendation['entry_points'].append(round(current_price * 1.01, 2))
            recommendation['stop_loss'] = round(current_price * 1.04, 2)
            recommendation['take_profit'] = round(current_price * 0.92, 2)
        else:
            recommendation['signal'] = 'ATTENDRE'
            recommendation['confidence'] = 50
            recommendation['analysis'].append("⏸️ Pas de signal clair - attendre un meilleur setup")
        return recommendation
    except Exception as e:
        return {'error': str(e)}


def calculate_rsi(prices, period=14):
    delta = prices.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    return rsi.iloc[-1] if not rsi.empty else 50

# ══════════════════════════════════════════════════════════════
# ROUTES — AUTH
# ══════════════════════════════════════════════════════════════

@app.route('/verify-token', methods=['GET', 'POST'])
def verify_token_page():
    email = request.args.get('email', session.get('pending_2fa_email', ''))
    if request.method == 'POST':
        entered_token = request.form.get('token', '').strip()
        stored_token  = session.get('pending_2fa_token', '')
        expires_str   = session.get('pending_2fa_expires', '')
        expired = False
        if expires_str:
            try:
                expires = datetime.fromisoformat(expires_str)
                expired = datetime.now() > expires
            except Exception:
                expired = True
        if expired:
            for k in list(session.keys()):
                if k.startswith('pending_2fa_'):
                    session.pop(k, None)
            return render_template('verify_token.html',
                email=email, error='Token expiré. Veuillez vous reconnecter.',
                token=None, message=None)
        if entered_token == stored_token:
            is_admin_login = session.pop('pending_2fa_is_admin_login', False)
            session['user_id']  = session.pop('pending_2fa_user_id',  None)
            session['username'] = session.pop('pending_2fa_username', '')
            session['email']    = session.pop('pending_2fa_email',    '')
            session['theme']    = session.pop('pending_2fa_theme',    'dark')
            session['role']     = session.pop('pending_2fa_role',     'user')
            session.pop('pending_2fa_token',   None)
            session.pop('pending_2fa_expires', None)
            session['admin_secondary_verified'] = False
            if is_admin_login:
                return redirect(url_for('admin_secondary_verify'))
            return redirect(url_for('dashboard'))
        else:
            return render_template('verify_token.html',
                email=email,
                token=session.get('pending_2fa_token'),
                error='Code incorrect. Vérifiez et réessayez.',
                message=None)
    pending_token = session.get('pending_2fa_token')
    if not pending_token:
        return redirect(url_for('login'))
    return render_template('verify_token.html',
        email=email,
        token=pending_token,
        error=None,
        message='Un code de vérification a été généré. Entrez-le ci-dessous pour continuer.')


@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        email = data.get('email')
        password = data.get('password')
        user = doc_to_dict(get_col('users').find_one({"email": email}))
        if user and check_password_hash(user['password'], password):
            token_2fa = str(random.randint(100000, 999999))
            session['pending_2fa_token']    = token_2fa
            session['pending_2fa_user_id']  = user['id']
            session['pending_2fa_username'] = user['username']
            session['pending_2fa_email']    = user['email']
            session['pending_2fa_theme']    = user.get('theme', 'dark')
            session['pending_2fa_role']     = user.get('role', 'user')
            session['pending_2fa_expires']  = (datetime.now() + timedelta(minutes=5)).isoformat()
            get_col('users').update_one({"id": user['id']}, {"$set": {"last_login": _now_iso()}})
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('verify_token_page', email=email)})
            return redirect(url_for('verify_token_page', email=email))
        if request.is_json:
            return jsonify({'success': False, 'message': 'Email ou mot de passe incorrect'}), 401
        return render_template('login.html', error='Email ou mot de passe incorrect')
    return render_template('login.html')


@app.route('/api/login-flyers', methods=['GET'])
def login_flyers():
    items = []
    try:
        journal_docs = docs_to_list(
            get_col('trading_journal').find(
                {"image_path": {"$ne": None, "$ne": ""}},
                sort=[("created_at", -1)],
                limit=20
            )
        )
        for row in journal_docs:
            img_url = row.get('image_path', '')
            if img_url and not img_url.startswith('/'):
                img_url = '/' + img_url
            pl = row.get('profit_loss', 0) or 0
            items.append({
                'type': 'trade',
                'img_url': img_url,
                'symbol': row.get('symbol', ''),
                'trade_type': row.get('type', ''),
                'profit_loss': pl,
                'strategy': row.get('strategy', ''),
                'date': row.get('date', ''),
                'entry_price': row.get('entry_price', 0),
                'exit_price': row.get('exit_price', 0),
                'badge': 'WIN' if pl > 0 else ('LOSS' if pl < 0 else 'BREAK'),
                'badge_color': '#00d4aa' if pl > 0 else ('#ff6b6b' if pl < 0 else '#ffd700')
            })
    except Exception:
        pass

    static_products = [
        {'name': 'Formation Débutant', 'spec': 'Bases du trading', 'price': '25 000 FCFA', 'promo': None,
         'img': 'formation_debutant.jpg', 'link': '/inscription-trading'},
        {'name': 'Formation Pro', 'spec': 'Mentoring personnalisé', 'price': '200 000 FCFA', 'promo': '-20%',
         'img': 'formation_pro.jpg', 'link': '/inscription-trading'},
    ]
    for p in static_products:
        flyer_path = os.path.join(app.static_folder, 'flyers', p['img'])
        if not os.path.exists(flyer_path):
            continue
        items.append({
            'type': 'product',
            'img_url': f"/static/flyers/{p['img']}",
            'name': p['name'],
            'spec': p['spec'],
            'price': p['price'],
            'promo': p['promo'],
            'link': p['link'],
            'badge': 'PROMO',
            'badge_color': '#00d4aa'
        })
    return jsonify({'success': True, 'items': items})


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        confirm_password = data.get('confirm_password')
        preferred_currency = data.get('preferred_currency', 'EUR')
        errors = []
        if not username or len(username) < 3:
            errors.append("Le nom d'utilisateur doit contenir au moins 3 caractères")
        if not email or '@' not in email:
            errors.append("Email invalide")
        if not password or len(password) < 6:
            errors.append("Le mot de passe doit contenir au moins 6 caractères")
        if password != confirm_password:
            errors.append("Les mots de passe ne correspondent pas")
        if errors:
            if request.is_json:
                return jsonify({'success': False, 'errors': errors}), 400
            return render_template('register.html', error=', '.join(errors))
        existing = get_col('users').find_one({"email": email})
        if existing:
            if request.is_json:
                return jsonify({'success': False, 'message': 'Cet email est déjà utilisé'}), 400
            return render_template('register.html', error='Cet email est déjà utilisé')
        try:
            uid = get_next_id("users")
            get_col('users').insert_one({
                "id": uid,
                "username": username,
                "email": email,
                "password": generate_password_hash(password),
                "preferred_currency": preferred_currency,
                "role": "user",
                "status": "active",
                "theme": "dark",
                "notifications_email": 1,
                "notifications_app": 1,
                "timezone": "Europe/Paris",
                "created_at": _now_iso(),
                "updated_at": _now_iso(),
                "last_login": None
            })
            session['user_id'] = uid
            session['username'] = username
            session['email'] = email
            session['theme'] = 'dark'
            session['role'] = 'user'
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            return redirect(url_for('dashboard'))
        except Exception as e:
            if request.is_json:
                return jsonify({'success': False, 'message': str(e)}), 500
            return render_template('register.html', error=f"Erreur lors de la création du compte: {str(e)}")
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ══════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════

@app.route('/dashboard')
@login_required
def dashboard():
    user_id = session['user_id']
    stats = {
        'net_worth': 0, 'monthly_cashflow': 0, 'expense_ratio': 0,
        'savings_rate': 0, 'total_revenue': 0, 'total_expenses': 0, 'trader_score': 0
    }
    recent_transactions = []
    unread_notifications = 0
    chart_labels = []; chart_revenus = []; chart_depenses = []; chart_solde = []
    donut_labels = []; donut_data = []
    recent_trainings = []; training_total = 0; training_total_min = 0; training_this_month = 0

    try:
        date_30 = _date_ago(30)
        date_6m = _date_ago(180)
        date_1m = datetime.now().replace(day=1).strftime('%Y-%m-%d')

        # Stats financières 30 jours
        pipeline = [
            {"$match": {"user_id": user_id, "date": {"$gte": date_30}}},
            {"$group": {
                "_id": None,
                "total_revenue": {"$sum": {"$cond": [{"$eq": ["$type", "revenue"]}, "$amount", 0]}},
                "total_expenses": {"$sum": {"$cond": [{"$eq": ["$type", "expense"]}, "$amount", 0]}}
            }}
        ]
        res = list(get_col('financial_transactions').aggregate(pipeline))
        if res:
            stats['total_revenue'] = res[0].get('total_revenue', 0) or 0
            stats['total_expenses'] = res[0].get('total_expenses', 0) or 0
            stats['monthly_cashflow'] = stats['total_revenue'] - stats['total_expenses']
            if stats['total_revenue'] > 0:
                stats['expense_ratio'] = (stats['total_expenses'] / stats['total_revenue']) * 100
                stats['savings_rate'] = 100 - stats['expense_ratio']

        # Valeur portfolio
        pipeline_port = [
            {"$match": {"user_id": user_id, "status": "open"}},
            {"$group": {"_id": None, "portfolio_value": {"$sum": {"$multiply": ["$quantity", "$current_price"]}}}}
        ]
        res_port = list(get_col('positions').aggregate(pipeline_port))
        portfolio_value = res_port[0]['portfolio_value'] if res_port else 0
        stats['net_worth'] = stats['monthly_cashflow'] + (portfolio_value or 0)

        # Dernier trader score
        last_score = doc_to_dict(get_col('trader_scores').find_one(
            {"user_id": user_id}, sort=[("created_at", -1)]
        ))
        if last_score:
            stats['trader_score'] = last_score['overall_score']

        # Transactions récentes
        recent_transactions = docs_to_list(
            get_col('transactions').find({"user_id": user_id}, sort=[("created_at", -1)], limit=10)
        )

        # Notifications non lues
        unread_notifications = get_col('notifications').count_documents(
            {"user_id": user_id, "is_read": 0}
        )

        # Graphique performance 6 mois
        perf_pipeline = [
            {"$match": {"user_id": user_id, "date": {"$gte": date_6m}}},
            {"$group": {
                "_id": {"$substr": ["$date", 0, 7]},
                "revenus": {"$sum": {"$cond": [{"$eq": ["$type", "revenue"]}, "$amount", 0]}},
                "depenses": {"$sum": {"$cond": [{"$eq": ["$type", "expense"]}, "$amount", 0]}}
            }},
            {"$sort": {"_id": 1}}
        ]
        perf_rows = list(get_col('financial_transactions').aggregate(perf_pipeline))
        month_names = {'01':'Jan','02':'Fév','03':'Mar','04':'Avr','05':'Mai',
                       '06':'Juin','07':'Juil','08':'Août','09':'Sep','10':'Oct','11':'Nov','12':'Déc'}
        chart_labels   = [month_names.get(r['_id'][-2:], r['_id']) for r in perf_rows]
        chart_revenus  = [round(r.get('revenus', 0) or 0, 2) for r in perf_rows]
        chart_depenses = [round(r.get('depenses', 0) or 0, 2) for r in perf_rows]
        chart_solde    = [round((r.get('revenus', 0) or 0) - (r.get('depenses', 0) or 0), 2) for r in perf_rows]

        # Donut catégories dépenses
        donut_pipeline = [
            {"$match": {"user_id": user_id, "type": "expense"}},
            {"$group": {"_id": "$category", "total": {"$sum": "$amount"}}},
            {"$sort": {"total": -1}},
            {"$limit": 6}
        ]
        cat_rows = list(get_col('financial_transactions').aggregate(donut_pipeline))
        donut_labels = [r['_id'] for r in cat_rows]
        donut_data   = [round(r.get('total', 0) or 0, 2) for r in cat_rows]

        # Formations récentes
        recent_trainings = docs_to_list(
            get_col('training_courses').find(
                {"user_id": user_id},
                sort=[("scheduled_date", -1), ("created_at", -1)],
                limit=5
            )
        )
        training_total = get_col('training_courses').count_documents({"user_id": user_id})
        tmins = list(get_col('training_courses').aggregate([
            {"$match": {"user_id": user_id}},
            {"$group": {"_id": None, "total": {"$sum": "$duration_minutes"}}}
        ]))
        training_total_min = tmins[0]['total'] if tmins else 0
        training_this_month = get_col('training_courses').count_documents(
            {"user_id": user_id, "scheduled_date": {"$gte": date_1m}}
        )
    except Exception as e:
        print(f"dashboard error: {e}")

    return render_template('dashboard.html',
        stats=stats,
        transactions=recent_transactions,
        unread_notifications=unread_notifications,
        user_role=session.get('role', 'user'),
        chart_labels=chart_labels,
        chart_revenus=chart_revenus,
        chart_depenses=chart_depenses,
        chart_solde=chart_solde,
        donut_labels=donut_labels,
        donut_data=donut_data,
        recent_trainings=recent_trainings,
        training_total=training_total,
        training_total_min=training_total_min,
        training_this_month=training_this_month
    )

# ══════════════════════════════════════════════════════════════
# FINANCES
# ══════════════════════════════════════════════════════════════

@app.route('/finances')
@login_required
def finances():
    user_id = session['user_id']
    filter_cat = request.args.get('category', '')
    filter_month = request.args.get('month', '')

    query = {"user_id": user_id}
    if filter_cat:
        query["category"] = filter_cat
    if filter_month:
        query["date"] = {"$regex": f"^{filter_month}"}

    transactions = docs_to_list(
        get_col('financial_transactions').find(query, sort=[("date", -1), ("time", -1)])
    )

    total_rev = sum(t['amount'] for t in transactions if t['type'] in ('revenue', 'receivable', 'credit'))
    total_exp = sum(t['amount'] for t in transactions if t['type'] in ('expense', 'debt'))
    balance = total_rev - total_exp
    savings_rate = max((balance / total_rev * 100) if total_rev > 0 else 0, 0)
    summary = {
        'total_revenue': total_rev,
        'total_expenses': total_exp,
        'net_balance': balance,
        'balance': balance,
        'savings_rate': savings_rate,
        'period': filter_month if filter_month else "Global"
    }

    # Chart 30 jours
    date_30 = _date_ago(30)
    chart_pipeline = [
        {"$match": {"user_id": user_id, "date": {"$gte": date_30}}},
        {"$group": {
            "_id": {"$substr": ["$date", 0, 10]},
            "rev": {"$sum": {"$cond": [{"$eq": ["$type", "revenue"]}, "$amount", 0]}},
            "exp": {"$sum": {"$cond": [{"$eq": ["$type", "expense"]}, "$amount", 0]}}
        }},
        {"$sort": {"_id": 1}}
    ]
    chart_raw = [{"day": r['_id'], "rev": r['rev'], "exp": r['exp']}
                 for r in get_col('financial_transactions').aggregate(chart_pipeline)]

    categories_cursor = get_col('financial_transactions').distinct("category", {"user_id": user_id})
    categories = list(categories_cursor)

    return render_template('finances.html',
        transactions=transactions,
        summary=summary,
        categories=categories,
        chart_data=json.dumps(chart_raw[::-1])
    )


@app.route('/api/add-transaction', methods=['POST'])
@login_required
def add_transaction():
    try:
        user_id  = session['user_id']
        t_type   = request.form.get('type')
        amount   = float(request.form.get('amount', 0))
        category = request.form.get('category')
        reason   = request.form.get('reason') or request.form.get('description') or ''
        t_date   = request.form.get('transaction_date') or request.form.get('date') or datetime.now().strftime('%Y-%m-%d')
        t_time   = request.form.get('time') or datetime.now().strftime('%H:%M:%S')
        currency        = request.form.get('currency', 'EUR')
        payment_method  = request.form.get('payment_method', '')
        tags            = ','.join(request.form.getlist('tags'))
        notes           = request.form.get('notes', '')
        emotional_state = ','.join(request.form.getlist('emotional_state'))
        energy_level    = request.form.get('energy_level', '3')
        if emotional_state:
            notes = f"{notes} [Émotions: {emotional_state}] [Énergie: {energy_level}/5]".strip()
        img_tag = ""
        if 'receipt_image' in request.files:
            file = request.files['receipt_image']
            if file and file.filename != '':
                filename = secure_filename(f"rec_{user_id}_{datetime.now().strftime('%m%d_%H%M%S')}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                img_tag = f" [IMG:{filename}]"
        tid = get_next_id("financial_transactions")
        get_col('financial_transactions').insert_one({
            "id": tid,
            "user_id": user_id,
            "type": t_type,
            "amount": amount,
            "category": category,
            "reason": f"{reason}{img_tag}",
            "date": t_date,
            "time": t_time,
            "status": 'Terminé',
            "currency": currency,
            "payment_method": payment_method,
            "tags": tags,
            "notes": notes,
            "created_at": _now_iso(),
            "updated_at": _now_iso()
        })
        flash('Transaction et justificatif enregistrés !', 'success')
    except Exception as e:
        flash(f'Erreur : {str(e)}', 'error')
    return redirect(url_for('finances'))


@app.route('/delete-transaction/<id>', methods=['POST'])
@login_required
def delete_transaction(id):
    get_col('financial_transactions').delete_one({"id": int(id), "user_id": session['user_id']})
    return redirect(url_for('finances'))


@app.route('/delete-journal/<id>', methods=['POST'])
@login_required
def delete_journal(id):
    get_col('trading_journal').delete_one({"id": int(id), "user_id": session['user_id']})
    return redirect(url_for('trading_journal'))


@app.route('/api/financial-transaction', methods=['POST'])
@login_required
def add_financial_transaction():
    data = request.get_json()
    user_id = session['user_id']
    required_fields = ['type', 'category', 'reason', 'amount', 'date', 'time']
    if not all(field in data for field in required_fields):
        return jsonify({'success': False, 'message': 'Champs requis manquants'}), 400
    try:
        tid = get_next_id("financial_transactions")
        get_col('financial_transactions').insert_one({
            "id": tid,
            "user_id": user_id,
            "type": data['type'],
            "category": data['category'],
            "subcategory": data.get('subcategory'),
            "reason": data['reason'],
            "usage": data.get('usage'),
            "amount": float(data['amount']),
            "currency": data.get('currency', 'EUR'),
            "date": data['date'],
            "time": data['time'],
            "payment_method": data.get('payment_method'),
            "reference": data.get('reference'),
            "status": data.get('status', 'completed'),
            "notes": data.get('notes'),
            "tags": data.get('tags'),
            "created_at": _now_iso(),
            "updated_at": _now_iso()
        })
        if float(data['amount']) > 1000:
            create_notification(user_id, 'info', 'Transaction importante',
                                f"Transaction de {data['amount']}€ enregistrée")
        return jsonify({'success': True, 'id': tid})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/delete-financial-transaction/<id>', methods=['DELETE', 'POST'])
@login_required
def delete_financial_transaction(id):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Suppression réservée aux administrateurs'}), 403
    try:
        get_col('financial_transactions').delete_one({"id": int(id), "user_id": session['user_id']})
        create_notification(session['user_id'], 'success', 'Transaction supprimée',
                            f'La transaction #{id} a été supprimée')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════
# EXPORT FINANCES
# ══════════════════════════════════════════════════════════════

@app.route('/api/export-finances')
@login_required
def export_finances():
    user_id = session['user_id']
    export_format = request.args.get('format', 'json')
    transactions = docs_to_list(
        get_col('financial_transactions').find({"user_id": user_id}, sort=[("date", -1)])
    )
    if export_format == 'json':
        return jsonify({'success': True, 'data': transactions})
    elif export_format in ('excel', 'csv'):
        df = pd.DataFrame(transactions)
        output = BytesIO()
        if export_format == 'excel':
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Finances')
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'finances_{datetime.now().strftime("%Y%m%d")}.xlsx')
        else:
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            return send_file(output, mimetype='text/csv', as_attachment=True,
                             download_name=f'finances_{datetime.now().strftime("%Y%m%d")}.csv')
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                textColor=colors.HexColor('#1a1a1a'), spaceAfter=30, alignment=1)
            elements.append(Paragraph('Transactions Financières', title_style))
            elements.append(Spacer(1, 20))
            date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10,
                textColor=colors.grey, alignment=1)
            elements.append(Paragraph(f'Généré le {datetime.now().strftime("%d/%m/%Y à %H:%M")}', date_style))
            elements.append(Spacer(1, 30))
            limited_transactions = transactions[:50]
            data = [['Date', 'Type', 'Catégorie', 'Raison', 'Montant']]
            total_revenue = 0; total_expense = 0
            for trans in limited_transactions:
                amount = float(trans['amount'])
                if trans['type'] == 'revenue':
                    total_revenue += amount; amount_str = f"+{amount:.2f}€"
                else:
                    total_expense += amount; amount_str = f"-{amount:.2f}€"
                data.append([trans['date'], trans['type'].capitalize(),
                              str(trans['category'])[:15], str(trans['reason'])[:20], amount_str])
            balance = total_revenue - total_expense
            data.append(['', '', '', 'SOLDE', f"{balance:.2f}€"])
            table = Table(data, colWidths=[1*inch, 1*inch, 1.2*inch, 1.5*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9)
            ]))
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            return send_file(output, mimetype='application/pdf', as_attachment=True,
                             download_name=f'finances_{datetime.now().strftime("%Y%m%d")}.pdf')
        except ImportError:
            return jsonify({'success': False, 'message': 'ReportLab non installé.'}), 500
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    return jsonify({'success': False, 'message': 'Format non supporté'}), 400

# ══════════════════════════════════════════════════════════════
# TRADING JOURNAL
# ══════════════════════════════════════════════════════════════

@app.route('/journal')
@login_required
def trading_journal():
    user_id = session['user_id']
    entries = docs_to_list(
        get_col('trading_journal').find({"user_id": user_id}, sort=[("date", -1), ("time", -1)])
    )
    return render_template('trading_journal.html', entries=entries)


@app.route('/api/journal-entry', methods=['POST'])
@login_required
def add_journal_entry():
    user_id = session['user_id']
    image_path = None
    if 'image' in request.files:
        file = request.files['image']
        if file and allowed_file(file.filename):
            filename = secure_filename(f"{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            image_path = filepath
    data = request.form if not request.is_json else request.get_json()
    try:
        jid = get_next_id("trading_journal")
        get_col('trading_journal').insert_one({
            "id": jid,
            "user_id": user_id,
            "symbol": data.get('symbol'),
            "date": data.get('date'),
            "time": data.get('time'),
            "type": data.get('type'),
            "quantity": float(data.get('quantity', 0)),
            "entry_price": float(data.get('entry_price', 0)),
            "exit_price": float(data.get('exit_price', 0)) if data.get('exit_price') else None,
            "profit_loss": float(data.get('profit_loss', 0)) if data.get('profit_loss') else None,
            "strategy": data.get('strategy'),
            "setup_description": data.get('setup_description'),
            "emotions": data.get('emotions'),
            "mistakes": data.get('mistakes'),
            "lessons_learned": data.get('lessons_learned'),
            "notes": data.get('notes'),
            "image_path": image_path,
            "market_conditions": data.get('market_conditions'),
            "risk_reward_ratio": float(data.get('risk_reward_ratio', 0)) if data.get('risk_reward_ratio') else None,
            "created_at": _now_iso()
        })
        if image_path:
            trade_data = {
                'profit_loss': float(data.get('profit_loss', 0)) if data.get('profit_loss') else None,
                'risk_reward_ratio': float(data.get('risk_reward_ratio', 0)) if data.get('risk_reward_ratio') else None,
                'strategy': data.get('strategy')
            }
            analysis = analyze_trade_image(image_path, trade_data)
            aid = get_next_id("ai_analysis")
            get_col('ai_analysis').insert_one({
                "id": aid,
                "user_id": user_id,
                "analysis_type": "trading",
                "subject": f"Journal Entry #{jid}",
                "insights": json.dumps(analysis),
                "created_at": _now_iso()
            })
        if request.is_json:
            return jsonify({'success': True, 'id': jid})
        return redirect(url_for('trading_journal'))
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/delete-journal-entry/<id>', methods=['POST'])
@login_required
def delete_journal_entry(id):
    if session.get('role') not in ('admin', 'superadmin'):
        flash('Suppression réservée aux administrateurs', 'danger')
        return redirect(url_for('trading_journal'))
    try:
        get_col('trading_journal').delete_one({"id": int(id), "user_id": session['user_id']})
        flash('Entrée supprimée', 'success')
        return redirect(url_for('trading_journal'))
    except Exception as e:
        flash(f'Erreur: {e}', 'danger')
        return redirect(url_for('trading_journal'))

# ══════════════════════════════════════════════════════════════
# TRADING
# ══════════════════════════════════════════════════════════════

@app.route('/trading')
@login_required
def trading():
    user_id = session['user_id']
    positions = docs_to_list(
        get_col('positions').find({"user_id": user_id, "status": "open"}, sort=[("created_at", -1)])
    )
    return render_template('trading.html', positions=positions)


@app.route('/api/execute-trade', methods=['POST'])
@login_required
def execute_trade():
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form.to_dict()
    user_id = session['user_id']
    if not data:
        return jsonify({'success': False, 'message': 'Données manquantes'}), 400
    required_fields = ['symbol', 'type', 'quantity', 'price']
    if not all(field in data for field in required_fields):
        return jsonify({'success': False, 'message': 'Champs requis manquants'}), 400
    symbol = data['symbol'].upper()
    trade_type = data['type']
    quantity = float(data['quantity'])
    price = float(data['price'])
    fees = float(data.get('fees', 0))
    strategy = data.get('strategy')
    amount = quantity * price
    if trade_type == 'sell':
        amount = amount - fees
    else:
        amount = -(amount + fees)
    try:
        tid = get_next_id("transactions")
        get_col('transactions').insert_one({
            "id": tid,
            "user_id": user_id,
            "symbol": symbol,
            "type": trade_type,
            "quantity": quantity,
            "price": price,
            "amount": amount,
            "fees": fees,
            "status": "completed",
            "strategy": strategy,
            "created_at": _now_iso()
        })
        if trade_type == 'buy':
            existing = doc_to_dict(get_col('positions').find_one(
                {"user_id": user_id, "symbol": symbol, "status": "open"}
            ))
            if existing:
                new_qty = existing['quantity'] + quantity
                new_avg = ((existing['quantity'] * existing['avg_price']) + (quantity * price)) / new_qty
                get_col('positions').update_one(
                    {"user_id": user_id, "symbol": symbol, "status": "open"},
                    {"$set": {"quantity": new_qty, "avg_price": new_avg, "updated_at": _now_iso()}}
                )
            else:
                pid = get_next_id("positions")
                get_col('positions').insert_one({
                    "id": pid,
                    "user_id": user_id,
                    "symbol": symbol,
                    "quantity": quantity,
                    "avg_price": price,
                    "current_price": price,
                    "status": "open",
                    "stop_loss": None,
                    "take_profit": None,
                    "created_at": _now_iso(),
                    "updated_at": _now_iso()
                })
        else:
            get_col('positions').update_one(
                {"user_id": user_id, "symbol": symbol, "status": "open"},
                {"$inc": {"quantity": -quantity}, "$set": {"updated_at": _now_iso()}}
            )
        return jsonify({'success': True, 'message': 'Transaction exécutée avec succès'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ══════════════════════════════════════════════════════════════
# PORTFOLIO
# ══════════════════════════════════════════════════════════════

@app.route('/portfolio')
@login_required
def portfolio():
    user_id = session['user_id']
    positions = docs_to_list(
        get_col('positions').find({"user_id": user_id, "status": "open"})
    )
    portfolio_stats = {'total_value': 0, 'total_cost': 0, 'total_pnl': 0,
                       'total_pnl_percent': 0, 'best_performer': None, 'worst_performer': None}
    for pos in positions:
        pos['market_value'] = pos['quantity'] * pos['current_price']
        pos['cost_basis']   = pos['quantity'] * pos['avg_price']
        pos['pnl']          = pos['market_value'] - pos['cost_basis']
        pos['pnl_percent']  = (pos['pnl'] / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0
        portfolio_stats['total_value'] += pos['market_value']
        portfolio_stats['total_cost']  += pos['cost_basis']
        portfolio_stats['total_pnl']   += pos['pnl']
    if portfolio_stats['total_cost'] > 0:
        portfolio_stats['total_pnl_percent'] = (portfolio_stats['total_pnl'] / portfolio_stats['total_cost'] * 100)
    if positions:
        positions_sorted = sorted(positions, key=lambda x: x['pnl_percent'], reverse=True)
        portfolio_stats['best_performer']  = positions_sorted[0]
        portfolio_stats['worst_performer'] = positions_sorted[-1]
    return render_template('portfolio.html', positions=positions, stats=portfolio_stats)


@app.route('/api/add-position', methods=['POST'])
@login_required
def add_position():
    data = request.get_json()
    user_id = session['user_id']
    required_fields = ['symbol', 'quantity', 'avg_price']
    if not all(field in data for field in required_fields):
        return jsonify({'success': False, 'message': 'Champs requis manquants'}), 400
    try:
        quantity  = float(data['quantity'])
        avg_price = float(data['avg_price'])
        if quantity <= 0 or avg_price <= 0:
            return jsonify({'success': False, 'message': 'La quantité et le prix doivent être positifs'}), 400
        current_price = avg_price
        try:
            ticker = yf.Ticker(data['symbol'])
            hist = ticker.history(period='1d')
            if not hist.empty:
                current_price = float(hist['Close'].iloc[-1])
        except:
            pass
        pid = get_next_id("positions")
        get_col('positions').insert_one({
            "id": pid,
            "user_id": user_id,
            "symbol": data['symbol'].upper(),
            "asset_type": data.get('asset_type', 'stock'),
            "quantity": quantity,
            "avg_price": avg_price,
            "current_price": current_price,
            "status": "open",
            "platform": data.get('platform', 'Manual'),
            "notes": data.get('notes', ''),
            "stop_loss": None,
            "take_profit": None,
            "created_at": _now_iso(),
            "updated_at": _now_iso()
        })
        return jsonify({'success': True, 'id': pid})
    except ValueError:
        return jsonify({'success': False, 'message': 'Valeurs numériques invalides'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/export-portfolio')
@login_required
def export_portfolio():
    user_id = session['user_id']
    export_format = request.args.get('format', 'json')
    positions = docs_to_list(
        get_col('positions').find({"user_id": user_id, "status": "open"}, sort=[("symbol", 1)])
    )
    for pos in positions:
        pos['market_value'] = pos['quantity'] * pos['current_price']
        pos['cost_basis']   = pos['quantity'] * pos['avg_price']
        pos['pnl']          = pos['market_value'] - pos['cost_basis']
        pos['pnl_percent']  = (pos['pnl'] / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0
    if export_format == 'json':
        return jsonify({'success': True, 'data': positions})
    elif export_format in ('excel', 'csv'):
        df = pd.DataFrame(positions)
        output = BytesIO()
        if export_format == 'excel':
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Portfolio')
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'portfolio_{datetime.now().strftime("%Y%m%d")}.xlsx')
        else:
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            return send_file(output, mimetype='text/csv', as_attachment=True,
                             download_name=f'portfolio_{datetime.now().strftime("%Y%m%d")}.csv')
    elif export_format == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                textColor=colors.HexColor('#1a1a1a'), spaceAfter=30, alignment=1)
            elements.append(Paragraph('Portfolio Report', title_style))
            elements.append(Spacer(1, 20))
            date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10,
                textColor=colors.grey, alignment=1)
            elements.append(Paragraph(f'Généré le {datetime.now().strftime("%d/%m/%Y à %H:%M")}', date_style))
            elements.append(Spacer(1, 30))
            data = [['Symbol', 'Quantity', 'Avg Price', 'Current Price', 'P&L', 'P&L %']]
            total_value = 0; total_cost = 0
            for pos in positions:
                data.append([pos['symbol'], f"{pos['quantity']:.2f}", f"{pos['avg_price']:.2f}€",
                              f"{pos['current_price']:.2f}€", f"{pos['pnl']:.2f}€", f"{pos['pnl_percent']:.2f}%"])
                total_value += pos['market_value']; total_cost += pos['cost_basis']
            total_pnl = total_value - total_cost
            total_pnl_percent = (total_pnl / total_cost * 100) if total_cost > 0 else 0
            data.append(['TOTAL', '', '', '', f"{total_pnl:.2f}€", f"{total_pnl_percent:.2f}%"])
            table = Table(data, colWidths=[1.2*inch, 1*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9)
            ]))
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            return send_file(output, mimetype='application/pdf', as_attachment=True,
                             download_name=f'portfolio_{datetime.now().strftime("%Y%m%d")}.pdf')
        except ImportError:
            return jsonify({'success': False, 'message': 'ReportLab non installé.'}), 500
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    return jsonify({'success': False, 'message': 'Format non supporté'}), 400


@app.route('/api/analyze-portfolio')
@login_required
def analyze_portfolio():
    user_id = session['user_id']
    positions = docs_to_list(get_col('positions').find({"user_id": user_id, "status": "open"}))
    if not positions:
        return jsonify({'success': True, 'analysis': "Aucune position dans le portfolio pour l'instant."})
    total_value = 0; total_cost = 0
    best_performer = None; worst_performer = None
    max_pnl_percent = float('-inf'); min_pnl_percent = float('inf')
    for pos in positions:
        market_value = pos['quantity'] * pos['current_price']
        cost_basis   = pos['quantity'] * pos['avg_price']
        pnl_percent  = ((market_value - cost_basis) / cost_basis * 100) if cost_basis > 0 else 0
        total_value += market_value; total_cost += cost_basis
        if pnl_percent > max_pnl_percent:
            max_pnl_percent = pnl_percent; best_performer = pos['symbol']
        if pnl_percent < min_pnl_percent:
            min_pnl_percent = pnl_percent; worst_performer = pos['symbol']
    total_pnl = total_value - total_cost
    total_pnl_percent = (total_pnl / total_cost * 100) if total_cost > 0 else 0
    analysis = f"""📊 Analyse de votre Portfolio\n\n💰 Valeur totale: {total_value:.2f}XAF\n📈 P&L total: {total_pnl:+.2f}XAF ({total_pnl_percent:+.2f}%)\n📦 Nombre de positions: {len(positions)}\n\n🌟 Meilleure performance: {best_performer} ({max_pnl_percent:+.2f}%)\n⚠️ Moins bonne performance: {worst_performer} ({min_pnl_percent:+.2f}%)\n\n💡 Recommandations:\n- {'Excellent rendement!' if total_pnl_percent > 10 else 'Continuez à diversifier votre portfolio'}\n- {'Pensez à prendre des bénéfices sur ' + best_performer if max_pnl_percent > 20 else 'Surveillez les opportunités de renforcement'}\n- {'Analysez ' + worst_performer + ' pour décider de conserver ou liquider' if min_pnl_percent < -10 else 'Portfolio bien équilibré'}"""
    return jsonify({'success': True, 'analysis': analysis,
                    'stats': {'total_value': total_value, 'total_pnl': total_pnl,
                              'total_pnl_percent': total_pnl_percent,
                              'positions_count': len(positions),
                              'best_performer': best_performer, 'worst_performer': worst_performer}})


@app.route('/api/delete-position/<id>', methods=['DELETE', 'POST'])
@login_required
def delete_position(id):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Suppression réservée aux administrateurs'}), 403
    try:
        get_col('positions').delete_one({"id": int(id), "user_id": session['user_id']})
        create_notification(session['user_id'], 'success', 'Position supprimée', f'La position #{id} a été supprimée')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/delete-trade/<id>', methods=['DELETE', 'POST'])
@login_required
def delete_trade(id):
    if session.get('role') not in ('admin', 'superadmin'):
        return jsonify({'success': False, 'error': 'Suppression réservée aux administrateurs'}), 403
    try:
        get_col('transactions').delete_one({"id": int(id), "user_id": session['user_id']})
        create_notification(session['user_id'], 'success', 'Trade supprimé', f'Le trade #{id} a été supprimé')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ══════════════════════════════════════════════════════════════
# AI ASSISTANT
# ══════════════════════════════════════════════════════════════

@app.route('/ai-assistant')
@login_required
def ai_assistant():
    return render_template('ai_assistant.html')


@app.route('/api/ai-chat', methods=['POST'])
@login_required
def ai_chat():
    data = request.get_json()
    user_id = session['user_id']
    question = data.get('question', '').lower()
    response = {'answer': '', 'data': None, 'charts': []}
    try:
        if 'pourquoi' in question and ('perdu' in question or 'perte' in question):
            date_30 = _date_ago(30)
            pipeline = [
                {"$match": {"user_id": user_id, "created_at": {"$gte": date_30}}},
                {"$group": {
                    "_id": "$symbol",
                    "total_sell": {"$sum": {"$cond": [{"$eq": ["$type", "sell"]}, "$amount", 0]}},
                    "total_buy":  {"$sum": {"$cond": [{"$eq": ["$type", "buy"]},  "$amount", 0]}},
                    "trade_count": {"$sum": 1}
                }},
                {"$match": {"$expr": {"$lt": [{"$add": ["$total_sell", "$total_buy"]}, 0]}}},
                {"$sort": {"total_sell": 1}}
            ]
            losing_trades = list(get_col('transactions').aggregate(pipeline))
            losing_trades = [{"symbol": r['_id'], "total_sell": r['total_sell'],
                               "total_buy": r['total_buy'], "trade_count": r['trade_count']} for r in losing_trades]
            if losing_trades:
                total_loss = sum(t['total_sell'] + t['total_buy'] for t in losing_trades)
                response['answer'] = f"Vous avez perdu {abs(total_loss):.2f}€ ce mois-ci. Les principales pertes: "
                response['answer'] += ", ".join([f"{t['symbol']} ({t['total_sell'] + t['total_buy']:.2f}€)" for t in losing_trades[:3]])
                response['data'] = losing_trades
            else:
                response['answer'] = "Vous n'avez pas enregistré de pertes ce mois-ci. Bravo!"
        elif 'stratégie' in question and ('rentable' in question or 'meilleur' in question):
            pipeline = [
                {"$match": {"user_id": user_id, "strategy": {"$ne": None}, "type": "sell"}},
                {"$group": {
                    "_id": "$strategy",
                    "trade_count": {"$sum": 1},
                    "total_profit": {"$sum": "$amount"},
                    "avg_profit": {"$avg": "$amount"},
                    "wins":   {"$sum": {"$cond": [{"$gt": ["$amount", 0]}, 1, 0]}},
                    "losses": {"$sum": {"$cond": [{"$lt": ["$amount", 0]}, 1, 0]}}
                }},
                {"$sort": {"total_profit": -1}}
            ]
            raw = list(get_col('transactions').aggregate(pipeline))
            strategies = [{"strategy": r['_id'], "trade_count": r['trade_count'],
                            "total_profit": r['total_profit'], "avg_profit": r['avg_profit'],
                            "wins": r['wins'], "losses": r['losses']} for r in raw]
            if strategies:
                best = strategies[0]
                win_rate = (best['wins'] / best['trade_count'] * 100) if best['trade_count'] > 0 else 0
                response['answer'] = (f"Votre meilleure stratégie est '{best['strategy']}' avec:\n"
                                      f"• Profit total: {best['total_profit']:.2f}€\n"
                                      f"• {best['trade_count']} trades\n"
                                      f"• Taux de réussite: {win_rate:.1f}%\n"
                                      f"• Profit moyen: {best['avg_profit']:.2f}€")
                response['data'] = strategies
            else:
                response['answer'] = "Vous n'avez pas encore de données de stratégie enregistrées."
        elif 'score' in question or 'performance' in question:
            score_data = calculate_trader_score(user_id)
            response['answer'] = (f"Votre score de trader est: {score_data['overall_score']:.1f}/100\n\n"
                                   f"Détails:\n• Rentabilité: {score_data['profitability_score']:.1f}/100\n"
                                   f"• Gestion du risque: {score_data['risk_management_score']:.1f}/100\n"
                                   f"• Discipline: {score_data['discipline_score']:.1f}/100\n"
                                   f"• Cohérence stratégique: {score_data['strategy_consistency_score']:.1f}/100\n"
                                   f"• Contrôle émotionnel: {score_data['emotional_control_score']:.1f}/100")
            if score_data['overall_score'] < 50:
                response['answer'] += "\n\n⚠️ Votre score est faible. Concentrez-vous sur la discipline et la gestion du risque."
            elif score_data['overall_score'] < 70:
                response['answer'] += "\n\n📈 Bon début ! Travaillez sur la cohérence de vos stratégies."
            else:
                response['answer'] += "\n\n✅ Excellent score ! Continuez ainsi!"
            response['data'] = score_data
        elif 'combien' in question and ('gagn' in question or 'perdu' in question):
            pipeline = [
                {"$match": {"user_id": user_id, "type": "sell"}},
                {"$group": {
                    "_id": None,
                    "total_gains":  {"$sum": {"$cond": [{"$gt": ["$amount", 0]}, "$amount", 0]}},
                    "total_losses": {"$sum": {"$cond": [{"$lt": ["$amount", 0]}, {"$abs": "$amount"}, 0]}},
                    "net_profit":   {"$sum": "$amount"}
                }}
            ]
            result = list(get_col('transactions').aggregate(pipeline))
            if result and result[0].get('total_gains'):
                r = result[0]
                response['answer'] = (f"Résultats de trading:\n• Gains totaux: {r['total_gains']:.2f}€\n"
                                       f"• Pertes totales: {r['total_losses']:.2f}€\n• Profit net: {r['net_profit']:.2f}€")
                response['answer'] += "\n\n✅ Vous êtes profitable!" if r['net_profit'] > 0 else "\n\n⚠️ Vous êtes en perte. Analysez vos trades."
            else:
                response['answer'] = "Vous n'avez pas encore de trades fermés."
        elif 'problème' in question or 'erreur' in question:
            patterns = analyze_trading_psychology(user_id)
            if patterns:
                response['answer'] = f"J'ai détecté {len(patterns)} problèmes:\n\n"
                for i, pattern in enumerate(patterns[:5], 1):
                    response['answer'] += f"{i}. {pattern['type'].upper()} ({pattern['severity']})\n   {pattern['description']}\n   💡 {pattern['recommendation']}\n\n"
                response['data'] = patterns
            else:
                response['answer'] = "Aucun problème majeur détecté. Continuez votre bon travail!"
        elif 'conseil' in question or 'recommandation' in question:
            patterns = analyze_trading_psychology(user_id)
            score_data = calculate_trader_score(user_id)
            response['answer'] = "Recommandations personnalisées:\n\n"
            if score_data['discipline_score'] < 60:
                response['answer'] += "1. 📋 Discipline: Créez un plan de trading et suivez-le strictement\n"
            if score_data['risk_management_score'] < 60:
                response['answer'] += "2. 🛡️ Risque: Utilisez toujours des stop-loss (max 2% par trade)\n"
            if score_data['emotional_control_score'] < 60:
                response['answer'] += "3. 🧠 Émotions: Tenez un journal de vos émotions avant chaque trade\n"
            if not response['answer'].strip().endswith('\n\n'):
                response['answer'] += "✅ Continuez votre excellent travail!"
        else:
            response['answer'] = "Je suis votre assistant trading IA. Posez-moi des questions comme:\n• 'Pourquoi ai-je perdu ce mois ?'\n• 'Quelle est ma meilleure stratégie ?'\n• 'Quel est mon score ?'\n• 'Combien ai-je gagné ?'"
    except Exception as e:
        response['answer'] = f"Erreur lors de l'analyse: {str(e)}"
    return jsonify(response)


@app.route('/api/ai-analyze-finances', methods=['GET', 'POST'])
@login_required
def ai_analyze_finances():
    user_id = session['user_id']
    date_30 = _date_ago(30)
    pipeline = [
        {"$match": {"user_id": user_id, "date": {"$gte": date_30}}},
        {"$group": {
            "_id": {"type": "$type", "category": "$category"},
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"total": -1}}
    ]
    categories = list(get_col('financial_transactions').aggregate(pipeline))

    totals_pipeline = [
        {"$match": {"user_id": user_id, "date": {"$gte": date_30}}},
        {"$group": {
            "_id": None,
            "total_revenue":  {"$sum": {"$cond": [{"$eq": ["$type", "revenue"]}, "$amount", 0]}},
            "total_expenses": {"$sum": {"$cond": [{"$eq": ["$type", "expense"]}, "$amount", 0]}}
        }}
    ]
    totals = list(get_col('financial_transactions').aggregate(totals_pipeline))
    total_revenue  = totals[0]['total_revenue']  if totals else 0
    total_expenses = totals[0]['total_expenses'] if totals else 0
    balance = total_revenue - total_expenses

    analysis = f"""📊 Analyse Financière des 30 derniers jours\n\n💰 Revenus: {total_revenue:.2f}XAF\n💸 Dépenses: {total_expenses:.2f}XAF\n📈 Solde: {balance:+.2f}XAF\n\n📋 Répartition par catégorie:\n"""
    for cat in categories[:5]:
        analysis += f"\n- {cat['_id']['category']}: {cat['total']:.2f}€ ({cat['count']} transactions)"
    analysis += f"\n\n💡 Recommandations:\n- {'Excellente gestion!' if balance > 0 else 'Attention aux dépenses'}\n- Taux d'épargne: {(balance/total_revenue*100 if total_revenue > 0 else 0):.1f}%"
    return jsonify({'success': True, 'analysis': analysis})

# ══════════════════════════════════════════════════════════════
# ANALYSIS
# ══════════════════════════════════════════════════════════════

@app.route('/analysis')
@login_required
def analysis():
    user_id = session['user_id']
    trader_score = calculate_trader_score(user_id)
    patterns = analyze_trading_psychology(user_id)
    recent_analyses = docs_to_list(
        get_col('ai_analysis').find({"user_id": user_id}, sort=[("created_at", -1)], limit=10)
    )
    return render_template('analysis.html', trader_score=trader_score,
                           patterns=patterns, analyses=recent_analyses)


@app.route('/api/analyze-finances', methods=['POST'])
@login_required
def analyze_finances():
    user_id = session['user_id']
    date_30 = _date_ago(30)
    pipeline = [
        {"$match": {"user_id": user_id, "date": {"$gte": date_30}}},
        {"$group": {
            "_id": None,
            "revenue":  {"$sum": {"$cond": [{"$eq": ["$type", "revenue"]}, "$amount", 0]}},
            "expenses": {"$sum": {"$cond": [{"$eq": ["$type", "expense"]}, "$amount", 0]}}
        }}
    ]
    result = list(get_col('financial_transactions').aggregate(pipeline))
    data = {"revenue": result[0]['revenue'] if result else 0,
            "expenses": result[0]['expenses'] if result else 0}
    insights = analyze_financial_report(data)
    aid = get_next_id("ai_analysis")
    get_col('ai_analysis').insert_one({
        "id": aid,
        "user_id": user_id,
        "analysis_type": "financial",
        "subject": "Monthly Report",
        "insights": json.dumps(insights),
        "created_at": _now_iso()
    })
    return jsonify(insights)


@app.route('/api/trading-recommendation/<symbol>')
@login_required
def get_trading_recommendation(symbol):
    recommendation = trading_recommendation(symbol.upper())
    return jsonify(recommendation)

# ══════════════════════════════════════════════════════════════
# SETTINGS
# ══════════════════════════════════════════════════════════════

@app.route('/settings')
@login_required
def settings():
    user_id = session['user_id']
    user = doc_to_dict(get_col('users').find_one({"id": user_id}))
    return render_template('settings.html', settings=user or {}, user_role=session.get('role', 'user'))


@app.route('/api/update-settings', methods=['POST'])
@login_required
def update_settings():
    data = request.get_json()
    user_id = session['user_id']
    try:
        updates = {}
        allowed_fields = ['username', 'email', 'preferred_currency', 'timezone',
                          'theme', 'notifications_email', 'notifications_app']
        for field in allowed_fields:
            if field in data:
                updates[field] = data[field]
        if 'password' in data and data['password']:
            updates['password'] = generate_password_hash(data['password'])
        if updates:
            updates['updated_at'] = _now_iso()
            get_col('users').update_one({"id": user_id}, {"$set": updates})
            if 'theme' in data:
                session['theme'] = data['theme']
        return jsonify({'success': True, 'message': 'Paramètres mis à jour'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ══════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════════════════════════════

@app.route('/notifications')
@login_required
def notifications():
    user_id = session['user_id']
    notifications_list = docs_to_list(
        get_col('notifications').find({"user_id": user_id}, sort=[("created_at", -1)], limit=50)
    )
    return render_template('notifications.html', notifications=notifications_list)


@app.route('/api/mark-notification-read/<id>', methods=['POST'])
@login_required
def mark_notification_read(id):
    get_col('notifications').update_one(
        {"id": int(id), "user_id": session['user_id']},
        {"$set": {"is_read": 1}}
    )
    return jsonify({'success': True})

# ══════════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════════

@app.route('/reports')
@login_required
def reports():
    user_id = session['user_id']
    reports_list = docs_to_list(
        get_col('reports').find({"user_id": user_id}, sort=[("created_at", -1)])
    )
    return render_template('reports.html', reports=reports_list)


@app.route('/api/generate-report', methods=['POST'])
@login_required
def generate_report():
    data = request.get_json()
    user_id = session['user_id']
    report_type  = data.get('type', 'monthly')
    period_start = data.get('start')
    period_end   = data.get('end')
    if not period_start or not period_end:
        today = datetime.now()
        if report_type == 'monthly':
            period_start = today.replace(day=1).strftime('%Y-%m-%d')
            period_end   = today.strftime('%Y-%m-%d')
        elif report_type == 'yearly':
            period_start = today.replace(month=1, day=1).strftime('%Y-%m-%d')
            period_end   = today.strftime('%Y-%m-%d')
        else:
            period_start = (today - timedelta(days=7)).strftime('%Y-%m-%d')
            period_end   = today.strftime('%Y-%m-%d')
    try:
        pipeline = [
            {"$match": {"user_id": user_id, "date": {"$gte": period_start, "$lte": period_end}}},
            {"$group": {
                "_id": None,
                "revenue":  {"$sum": {"$cond": [{"$eq": ["$type", "revenue"]}, "$amount", 0]}},
                "expenses": {"$sum": {"$cond": [{"$eq": ["$type", "expense"]}, "$amount", 0]}}
            }}
        ]
        result = list(get_col('financial_transactions').aggregate(pipeline))
        revenue  = result[0]['revenue']  if result else 0
        expenses = result[0]['expenses'] if result else 0
        profit   = revenue - expenses
        profit_margin = (profit / revenue * 100) if revenue > 0 else 0
        rid = get_next_id("reports")
        get_col('reports').insert_one({
            "id": rid,
            "user_id": user_id,
            "title": f"Rapport {report_type} - {period_start} à {period_end}",
            "report_type": report_type,
            "period_start": period_start,
            "period_end": period_end,
            "revenue": revenue,
            "expenses": expenses,
            "profit": profit,
            "profit_margin": profit_margin,
            "created_at": _now_iso(),
            "updated_at": _now_iso()
        })
        return jsonify({'success': True, 'report_id': rid})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/download-report/<id>', methods=['GET'])
@login_required
def download_report(id):
    user_id = session['user_id']
    report = doc_to_dict(get_col('reports').find_one({"id": int(id), "user_id": user_id}))
    if not report:
        return jsonify({'error': 'Rapport introuvable'}), 404
    try:
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        buffer = BytesIO()
        c = pdf_canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        c.setFillColor(colors.HexColor('#00d4aa'))
        c.rect(0, height-80, width, 80, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 22)
        c.drawCentredString(width/2, height-45, "KENGNI FINANCE")
        c.setFont("Helvetica", 11)
        c.drawCentredString(width/2, height-65, "k-ni chez Htech-training | Rapport Certifié")
        c.setFillColor(colors.HexColor('#1a1a2e'))
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, height-115, f"Rapport: {report.get('title', 'N/A')}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height-140, f"Période: {report.get('period_start','N/A')} → {report.get('period_end','N/A')}")
        c.drawString(50, height-160, f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
        c.setStrokeColor(colors.HexColor('#00d4aa'))
        c.setLineWidth(2)
        c.line(50, height-175, width-50, height-175)
        y = height-220
        revenue  = float(report.get('revenue') or 0)
        expenses = float(report.get('expenses') or 0)
        profit   = float(report.get('profit') or revenue - expenses)
        margin   = float(report.get('profit_margin') or (profit/revenue*100 if revenue > 0 else 0))
        rows = [
            ("💰 Revenus Total",      f"{revenue:,.2f} €",  '#00c853'),
            ("💸 Dépenses Total",     f"{expenses:,.2f} €", '#d50000'),
            ("📈 Profit / Perte",     f"{profit:+,.2f} €",  '#00c853' if profit >= 0 else '#d50000'),
            ("📊 Marge Bénéficiaire", f"{margin:.1f} %",    '#1565c0'),
        ]
        for label, value, color in rows:
            c.setFillColor(colors.HexColor('#f5f5f5'))
            c.rect(50, y-8, width-100, 30, fill=True, stroke=False)
            c.setFillColor(colors.HexColor('#333333'))
            c.setFont("Helvetica-Bold", 12)
            c.drawString(65, y+8, label)
            c.setFillColor(colors.HexColor(color))
            c.setFont("Helvetica-Bold", 13)
            c.drawRightString(width-65, y+8, value)
            y -= 42
        c.setFillColor(colors.HexColor('#f0f0f0'))
        c.rect(0, 0, width, 45, fill=True, stroke=False)
        c.setFillColor(colors.HexColor('#888888'))
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, 28, "Document certifié — Kengni Finance v2.1 — © 2025 Tous droits réservés")
        c.drawCentredString(width/2, 14, "k-ni chez Htech-training")
        c.save()
        buffer.seek(0)
        filename = f"rapport_{report.get('report_type','custom')}_{report.get('period_start','').replace('-','')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': f'Erreur PDF: {str(e)}'}), 500


@app.route('/report/<id>')
@login_required
def view_report(id):
    return redirect(url_for('download_report', id=id))

# ══════════════════════════════════════════════════════════════
# HISTORY
# ══════════════════════════════════════════════════════════════

@app.route('/history')
@login_required
def history():
    user_id = session['user_id']
    transactions = docs_to_list(
        get_col('transactions').find({"user_id": user_id}, sort=[("created_at", -1)], limit=200)
    )
    return render_template('history.html', transactions=transactions)

# ══════════════════════════════════════════════════════════════
# IMAGE SPAM MANAGER
# ══════════════════════════════════════════════════════════════

@app.route('/image-spam')
@login_required
def image_spam():
    return render_template('image_spam_manager.html')

# ══════════════════════════════════════════════════════════════
# ADMIN PANEL
# ══════════════════════════════════════════════════════════════

@app.route(f'/{ADMIN_SECRET_TOKEN}')
def admin_secret_entry():
    if 'user_id' in session and session.get('role') in ('admin', 'superadmin'):
        return redirect(url_for('admin_panel'))
    return render_template('admin_login.html', token=ADMIN_SECRET_TOKEN)


@app.route(f'/{ADMIN_SECRET_TOKEN}/auth', methods=['POST'])
def admin_auth():
    data = request.get_json() if request.is_json else request.form
    email    = data.get('email', '').strip()
    password = data.get('password', '').strip()
    user = doc_to_dict(get_col('users').find_one({"email": email, "role": {"$in": ['admin', 'superadmin']}}))
    if user and check_password_hash(user['password'], password):
        token_2fa = str(random.randint(100000, 999999))
        session['pending_2fa_token']    = token_2fa
        session['pending_2fa_user_id']  = user['id']
        session['pending_2fa_username'] = user['username']
        session['pending_2fa_email']    = user['email']
        session['pending_2fa_theme']    = user.get('theme', 'dark')
        session['pending_2fa_role']     = user['role']
        session['pending_2fa_expires']  = (datetime.now() + timedelta(minutes=5)).isoformat()
        session['pending_2fa_is_admin_login'] = True
        get_col('users').update_one({"id": user['id']}, {"$set": {"last_login": _now_iso()}})
        if request.is_json:
            return jsonify({'success': True, 'redirect': url_for('verify_token_page', email=user['email'])})
        return redirect(url_for('verify_token_page', email=user['email']))
    if request.is_json:
        return jsonify({'success': False, 'message': 'Identifiants incorrects'}), 401
    from flask import abort; abort(404)


@app.route('/admin/secondary-verify', methods=['GET', 'POST'])
@admin_required
def admin_secondary_verify():
    error = None
    if request.method == 'POST':
        pwd = (request.get_json() or request.form).get('secondary_password', '')
        session['admin_sec_attempts'] = session.get('admin_sec_attempts', 0) + 1
        if session['admin_sec_attempts'] > 3:
            session.clear()
            if request.is_json:
                return jsonify({'success': False, 'message': 'Trop de tentatives — déconnexion'}), 403
            flash('Trop de tentatives incorrectes. Session terminée.', 'danger')
            return redirect(url_for('login'))
        if pwd == ADMIN_SECONDARY_PASSWORD:
            session['admin_secondary_verified'] = True
            session['admin_sec_attempts'] = 0
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('admin_panel')})
            return redirect(url_for('admin_panel'))
        else:
            remaining = 3 - session['admin_sec_attempts']
            error = f'Mot de passe secondaire incorrect. {remaining} tentative(s) restante(s).'
            if request.is_json:
                return jsonify({'success': False, 'message': error}), 401
    return render_template('admin_secondary.html', error=error)


@app.route('/admin')
@admin_required
def admin_panel():
    if not session.get('admin_secondary_verified'):
        return redirect(url_for('admin_secondary_verify'))
    users = docs_to_list(get_col('users').find({}, sort=[("created_at", -1)],
                                                projection={"password": 0}))
    stats = {
        'total_users':        get_col('users').count_documents({}),
        'active_users':       get_col('users').count_documents({"status": "active"}),
        'admins':             get_col('users').count_documents({"role": {"$in": ['admin', 'superadmin']}}),
        'total_transactions': get_col('financial_transactions').count_documents({})
    }
    return render_template('admin.html', users=users, stats=stats,
                           current_role=session.get('role'), token=ADMIN_SECRET_TOKEN)


@app.route('/admin/create-user', methods=['POST'])
@admin_required
def admin_create_user():
    data = request.get_json()
    username, email, password = data.get('username','').strip(), data.get('email','').strip(), data.get('password','').strip()
    role, status = data.get('role', 'user'), data.get('status', 'active')
    allowed = ['viewer', 'user', 'editor', 'admin']
    if session.get('role') == 'superadmin': allowed.append('superadmin')
    if not all([username, email, password]):
        return jsonify({'success': False, 'message': 'Tous les champs sont requis'}), 400
    if role not in allowed:
        return jsonify({'success': False, 'message': 'Rôle non autorisé'}), 403
    if get_col('users').find_one({"email": email}):
        return jsonify({'success': False, 'message': 'Email déjà utilisé'}), 409
    uid = get_next_id("users")
    get_col('users').insert_one({
        "id": uid,
        "username": username,
        "email": email,
        "password": generate_password_hash(password),
        "role": role,
        "status": status,
        "preferred_currency": "EUR",
        "timezone": "Europe/Paris",
        "theme": "dark",
        "notifications_email": 1,
        "notifications_app": 1,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
        "last_login": None
    })
    return jsonify({'success': True, 'message': f'Compte créé (ID {uid})', 'id': uid})


@app.route('/admin/update-user/<user_id>', methods=['POST'])
@admin_required
def admin_update_user(user_id):
    data = request.get_json()
    role, status = data.get('role'), data.get('status')
    allowed = ['viewer', 'user', 'editor', 'admin']
    if session.get('role') == 'superadmin': allowed.append('superadmin')
    if role and role not in allowed:
        return jsonify({'success': False, 'message': 'Rôle non autorisé'}), 403
    updates = {"updated_at": _now_iso()}
    if role:   updates['role']   = role
    if status: updates['status'] = status
    get_col('users').update_one({"id": int(user_id)}, {"$set": updates})
    return jsonify({'success': True, 'message': 'Utilisateur mis à jour'})


@app.route('/admin/reset-password/<user_id>', methods=['POST'])
@admin_required
def admin_reset_password(user_id):
    data = request.get_json()
    password = data.get('password', '').strip()
    if len(password) < 6:
        return jsonify({'success': False, 'message': 'Mot de passe trop court'}), 400
    get_col('users').update_one(
        {"id": int(user_id)},
        {"$set": {"password": generate_password_hash(password), "updated_at": _now_iso()}}
    )
    return jsonify({'success': True, 'message': 'Mot de passe réinitialisé'})


@app.route('/admin/delete-user/<user_id>', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    if int(user_id) == session['user_id']:
        return jsonify({'success': False, 'message': 'Impossible de supprimer votre propre compte'}), 400
    get_col('users').delete_one({"id": int(user_id)})
    return jsonify({'success': True, 'message': 'Utilisateur supprimé'})


@app.route('/api/admin/users')
@admin_required
def api_admin_users():
    users = docs_to_list(get_col('users').find(
        {}, sort=[("username", 1)],
        projection={"password": 0}
    ))
    return jsonify({'success': True, 'users': users})

# ══════════════════════════════════════════════════════════════
# TRAINING — Gestion des cours
# ══════════════════════════════════════════════════════════════

def detect_thumbnail(url):
    if not url:
        return '/static/img/course_default.svg'
    url_lower = url.lower()
    if 'claude.ai' in url_lower:
        return '/static/img/claude_thumb.svg'
    if 'chat.openai.com' in url_lower or 'chatgpt.com' in url_lower:
        return '/static/img/chatgpt_thumb.svg'
    yt = re.search(r'(?:v=|youtu\.be/)([A-Za-z0-9_-]{11})', url)
    if yt:
        return f'https://img.youtube.com/vi/{yt.group(1)}/hqdefault.jpg'
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            html = resp.read(30000).decode('utf-8', errors='ignore')
        og = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\'](.*?)["\']', html)
        if og: return og.group(1)
        og2 = re.search(r'<meta[^>]+content=["\'](.*?)["\'][^>]+property=["\']og:image["\']', html)
        if og2: return og2.group(1)
    except Exception:
        pass
    return '/static/img/course_default.svg'


@app.route('/training')
@login_required
def training():
    user_id = session['user_id']
    courses = docs_to_list(
        get_col('training_courses').find({"user_id": user_id}, sort=[("day_of_week", 1), ("created_at", -1)])
    )
    days_order = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche', 'Non défini']
    courses_by_day = {d: [] for d in days_order}
    for c in courses:
        day = c.get('day_of_week', 'Non défini')
        if day not in days_order: day = 'Non défini'
        try:
            c['position_images'] = json.loads(c.get('position_images') or '[]')
        except Exception:
            c['position_images'] = []
        courses_by_day[day].append(c)
    stats = {
        'total':          len(courses),
        'published':      sum(1 for c in courses if c['is_published']),
        'scheduled':      sum(1 for c in courses if c.get('scheduled_date')),
        'total_duration': sum((c['duration_minutes'] or 0) for c in courses),
    }
    return render_template('training.html', courses_by_day=courses_by_day, stats=stats)


@app.route('/training/add', methods=['POST'])
@login_required
def training_add():
    url = request.form.get('course_url', '').strip()
    thumbnail = request.form.get('thumbnail_url', '').strip() or detect_thumbnail(url)
    position_images = []
    for key in sorted(request.files.keys()):
        if key.startswith('position_img_') and not key.endswith('_caption'):
            file = request.files[key]
            if file and file.filename:
                fname = secure_filename(f"pos_{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                file.save(fpath)
                position_images.append(f'/static/uploads/{fname}')
    cid = get_next_id("training_courses")
    get_col('training_courses').insert_one({
        "id": cid,
        "user_id": session['user_id'],
        "title": request.form.get('title', 'Sans titre'),
        "description": request.form.get('description', ''),
        "course_url": url,
        "thumbnail_url": thumbnail,
        "category": request.form.get('category', 'Général'),
        "level": request.form.get('level', 'debutant'),
        "day_of_week": request.form.get('day_of_week', 'Non défini'),
        "scheduled_date": request.form.get('scheduled_date', ''),
        "duration_minutes": int(request.form.get('duration_minutes', 0) or 0),
        "tags": request.form.get('tags', ''),
        "is_published": 1 if request.form.get('is_published') else 0,
        "participant_names": request.form.get('participants', ''),
        "analyses": request.form.get('analyses', ''),
        "strategies": request.form.get('strategies', ''),
        "position_images": json.dumps(position_images),
        "time_start": request.form.get('time_start', ''),
        "time_end": request.form.get('time_end', ''),
        "view_count": 0,
        "created_at": _now_iso(),
        "updated_at": _now_iso()
    })
    return jsonify({'success': True, 'id': cid, 'thumbnail': thumbnail})


@app.route('/training/update/<cid>', methods=['POST'])
@login_required
def training_update(cid):
    url = request.form.get('course_url', '').strip()
    thumbnail = request.form.get('thumbnail_url', '').strip() or detect_thumbnail(url)
    existing_doc = doc_to_dict(get_col('training_courses').find_one(
        {"id": int(cid), "user_id": session['user_id']}
    ))
    try:
        existing_images = json.loads(existing_doc.get('position_images') if existing_doc else '[]') or []
    except Exception:
        existing_images = []
    imgs_to_delete = request.form.getlist('delete_images')
    existing_images = [img for img in existing_images if img not in imgs_to_delete]
    for key in sorted(request.files.keys()):
        if key.startswith('position_img_') and not key.endswith('_caption'):
            file = request.files[key]
            if file and file.filename:
                fname = secure_filename(f"pos_{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                file.save(fpath)
                existing_images.append(f'/static/uploads/{fname}')
    get_col('training_courses').update_one(
        {"id": int(cid), "user_id": session['user_id']},
        {"$set": {
            "title": request.form.get('title'),
            "description": request.form.get('description'),
            "course_url": url,
            "thumbnail_url": thumbnail,
            "category": request.form.get('category'),
            "level": request.form.get('level'),
            "day_of_week": request.form.get('day_of_week'),
            "scheduled_date": request.form.get('scheduled_date'),
            "duration_minutes": int(request.form.get('duration_minutes', 0) or 0),
            "tags": request.form.get('tags'),
            "is_published": 1 if request.form.get('is_published') else 0,
            "participant_names": request.form.get('participants', ''),
            "analyses": request.form.get('analyses', ''),
            "strategies": request.form.get('strategies', ''),
            "position_images": json.dumps(existing_images),
            "time_start": request.form.get('time_start', ''),
            "time_end": request.form.get('time_end', ''),
            "updated_at": _now_iso()
        }}
    )
    return jsonify({'success': True, 'thumbnail': thumbnail})


@app.route('/training/delete/<cid>', methods=['POST', 'DELETE'])
@login_required
def training_delete(cid):
    get_col('training_courses').delete_one({"id": int(cid), "user_id": session['user_id']})
    return jsonify({'success': True})


@app.route('/api/training/fetch-thumb', methods=['POST'])
@login_required
def training_fetch_thumb():
    data = request.get_json() or {}
    url = data.get('url', '')
    thumb = detect_thumbnail(url)
    return jsonify({'thumbnail': thumb})

# ══════════════════════════════════════════════════════════════
# KENGNI TRADING ACADEMY — Inscriptions & Leads
# ══════════════════════════════════════════════════════════════

@app.route('/inscription-trading', methods=['GET'])
def training_registration():
    success = request.args.get('success')
    wa      = request.args.get('wa', '')
    return render_template('inscription_trading.html', success=success, wa=wa)


@app.route('/inscription-trading', methods=['POST'])
def register_trading_lead():
    from urllib.parse import quote
    full_name      = request.form.get('full_name', '').strip()
    email          = request.form.get('email', '').strip().lower()
    whatsapp       = request.form.get('whatsapp', '').strip()
    level_selected = request.form.get('level_selected', '').strip()
    capital        = request.form.get('capital', '').strip()
    objective      = request.form.get('objective', '').strip()
    source         = request.form.get('source', 'Non renseigné').strip()
    errors = []
    if not full_name or len(full_name) < 2:   errors.append("Le nom complet est requis.")
    if not email or '@' not in email:          errors.append("Adresse email invalide.")
    if not whatsapp or len(whatsapp.replace(' ', '')) < 8: errors.append("Numéro WhatsApp requis.")
    if not level_selected: errors.append("Veuillez choisir un niveau de formation.")
    if errors:
        for err in errors: flash(err, 'error')
        return redirect(url_for('training_registration'))
    # Vérification doublon
    if get_col('training_leads').find_one({"email": email, "level_selected": level_selected}):
        flash(f"Vous êtes déjà inscrit(e) à la formation {level_selected} avec cet email.", 'error')
        return redirect(url_for('training_registration'))
    lid = get_next_id("training_leads")
    get_col('training_leads').insert_one({
        "id": lid,
        "full_name": full_name,
        "email": email,
        "whatsapp": whatsapp,
        "level_selected": level_selected,
        "capital": capital or None,
        "objective": objective or None,
        "source": source,
        "status": "Nouveau",
        "notes": None,
        "user_id": session.get('user_id'),
        "payment_method": None,
        "payment_ref": None,
        "payment_status": "En attente",
        "amount_paid": 0,
        "sincire_sent_at": None,
        "created_at": _now_iso()
    })
    flash("Inscription confirmée ! Nous vous contacterons sur WhatsApp très bientôt. 🎉", 'success')
    return redirect(url_for('training_registration', success=1, wa=whatsapp))


@app.route('/admin/leads')
@login_required
@admin_required
def admin_leads():
    leads = docs_to_list(get_col('training_leads').find({}, sort=[("created_at", -1)]))
    stats = {'total': 0, 'nouveau': 0, 'contacte': 0, 'inscrit': 0, 'paye': 0}
    stats['total']    = len(leads)
    stats['nouveau']  = sum(1 for l in leads if l['status'] == 'Nouveau')
    stats['contacte'] = sum(1 for l in leads if l['status'] == 'Contacté')
    stats['inscrit']  = sum(1 for l in leads if l['status'] == 'Inscrit')
    stats['paye']     = sum(1 for l in leads if l['status'] == 'Payé')
    return render_template('admin_leads.html', leads=leads, stats=stats)


@app.route('/admin/leads/<lead_id>/status', methods=['POST'])
@login_required
@admin_required
def update_lead_status(lead_id):
    new_status = request.form.get('status', '').strip()
    if new_status not in ['Nouveau', 'Contacté', 'Inscrit', 'Payé']:
        flash("Statut invalide.", 'error')
        return redirect(url_for('admin_leads'))
    get_col('training_leads').update_one({"id": int(lead_id)}, {"$set": {"status": new_status}})
    flash(f"Statut mis à jour : {new_status}", 'success')
    return redirect(url_for('admin_leads'))


@app.route('/admin/leads/<lead_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_lead(lead_id):
    get_col('training_leads').delete_one({"id": int(lead_id)})
    flash("Lead supprimé.", 'success')
    return redirect(url_for('admin_leads'))


def _send_sincire_email(lead: dict) -> bool:
    cfg      = GMAIL_CONFIG
    level    = lead.get('level_selected', 'Formation')
    name     = lead.get('full_name', 'Cher(e) prospect(e)')
    prospect_email = lead.get('email', '')
    prices   = FORMATION_PRICES.get(level, {'xaf': 50000, 'eur': 76})
    pay      = PAYMENT_INFO
    if not prospect_email:
        return False
    html = f'''<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"/></head>
<body style="margin:0;padding:0;background:#0a0f1a;font-family:'Segoe UI',Arial,sans-serif;">
<div style="max-width:600px;margin:0 auto;padding:24px;">
  <div style="background:linear-gradient(135deg,#0d1b2a,#1a2a3a);border-radius:18px 18px 0 0;padding:36px 32px;text-align:center;border-bottom:3px solid #00d4aa;">
    <div style="font-size:3rem;margin-bottom:10px;">🎓</div>
    <h1 style="color:#fff;margin:0;font-size:22px;font-weight:800;">Kengni Trading Academy</h1>
    <p style="color:#00d4aa;margin:8px 0 0;font-size:14px;font-weight:600;">Votre place est réservée — Finalisez votre inscription</p>
  </div>
  <div style="background:#111827;padding:32px;border-radius:0 0 18px 18px;border:1px solid #1e2a3a;border-top:none;">
    <p style="color:#e0e0e0;font-size:15px;line-height:1.7;margin:0 0 20px;">
      Bonjour <strong style="color:#00d4aa;">{name}</strong>,<br><br>
      Merci pour votre intérêt pour la formation <strong style="color:#fff;">"{level}"</strong> !
    </p>
    <div style="background:linear-gradient(135deg,rgba(0,212,170,.15),rgba(0,212,170,.05));border:1px solid rgba(0,212,170,.3);border-radius:12px;padding:20px;text-align:center;margin-bottom:24px;">
      <div style="font-size:.8rem;color:#888;margin-bottom:6px;">Montant à régler — {level}</div>
      <div style="font-size:2rem;font-weight:800;color:#00d4aa;">{prices['xaf']:,} FCFA</div>
      <div style="font-size:.9rem;color:#888;margin-top:4px;">≈ {prices['eur']} EUR</div>
    </div>
    <div style="background:#0d1b2a;border-radius:12px;padding:16px 18px;margin-bottom:10px;">
      <div style="color:#ff6b00;font-weight:700;">Orange Money: {pay['orange_money']['numero']}</div>
      <div style="color:#ffd700;font-weight:700;">MTN MoMo: {pay['mtn_money']['numero']}</div>
      <div style="color:#009cde;font-weight:700;">PayPal: {pay['paypal']['adresse']}</div>
    </div>
    <div style="border-top:1px solid #1e2a3a;padding-top:16px;text-align:center;margin-top:8px;">
      <p style="color:#444;font-size:11px;margin:0;">Kengni Trading Academy · fabrice.kengni12@gmail.com</p>
    </div>
  </div>
</div>
</body></html>'''
    text = (f"Bonjour {name},\n\nVotre place pour la formation \"{level}\" est réservée !\n\n"
            f"MONTANT : {prices['xaf']:,} FCFA (≈ {prices['eur']} EUR)\n\n"
            f"PAIEMENT :\n• Orange Money : {pay['orange_money']['numero']}\n"
            f"• MTN MoMo : {pay['mtn_money']['numero']}\n"
            f"• PayPal / Crypto : {pay['paypal']['adresse']}\n\n"
            f"Après paiement, envoyez la capture sur WhatsApp : +237 695 072 759\n\n— Kengni Trading Academy")
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🎓 Finalisez votre inscription — {level} | Kengni Trading Academy"
        msg['From']    = f"Kengni Trading Academy <{cfg['sender_email']}>"
        msg['To']      = prospect_email
        msg['Reply-To'] = cfg['sender_email']
        msg.attach(MIMEText(text, 'plain', 'utf-8'))
        msg.attach(MIMEText(html, 'html', 'utf-8'))
        with smtplib.SMTP(cfg['smtp_host'], cfg['smtp_port']) as s:
            s.ehlo(); s.starttls()
            s.login(cfg['sender_email'], cfg['smtp_password'])
            s.sendmail(cfg['sender_email'], prospect_email, msg.as_string())
        print(f"[Sincire] ✅ Email envoyé à {prospect_email}")
        return True
    except Exception as e:
        print(f"[Sincire] ❌ Erreur : {e}")
        return False


@app.route('/admin/leads/<lead_id>/sincire', methods=['POST'])
@login_required
@admin_required
def sincire_lead(lead_id):
    lead = doc_to_dict(get_col('training_leads').find_one({"id": int(lead_id)}))
    if not lead:
        return jsonify({'success': False, 'error': 'Lead introuvable'}), 404
    ok = _send_sincire_email(lead)
    if ok:
        now_iso = _now_iso()
        get_col('training_leads').update_one(
            {"id": int(lead_id)},
            {"$set": {"sincire_sent_at": now_iso, "status": "Contacté"}}
        )
        return jsonify({'success': True, 'message': f"✅ Email sincire envoyé à {lead.get('email')} !",
                        'sent_at': now_iso[:16].replace('T', ' à ')})
    return jsonify({'success': False, 'error': "Échec de l'envoi — vérifiez la config Gmail."}), 500


@app.route('/admin/leads/<lead_id>/update-payment', methods=['POST'])
@login_required
@admin_required
def update_lead_payment(lead_id):
    data = request.get_json() or {}
    get_col('training_leads').update_one(
        {"id": int(lead_id)},
        {"$set": {
            "payment_method": data.get('payment_method', ''),
            "payment_ref":    data.get('payment_ref', ''),
            "payment_status": data.get('payment_status', 'En attente'),
            "amount_paid":    float(data.get('amount_paid', 0) or 0)
        }}
    )
    return jsonify({'success': True})


@app.route('/api/export-leads')
@login_required
@admin_required
def export_leads():
    fmt   = request.args.get('format', 'csv').lower()
    leads = docs_to_list(get_col('training_leads').find({}, sort=[("created_at", -1)]))
    data = [{
        'ID':        l['id'],
        'Nom':       l['full_name'],
        'Email':     l['email'],
        'WhatsApp':  l['whatsapp'],
        'Formation': l['level_selected'],
        'Capital':   l.get('capital') or '',
        'Objectif':  l.get('objective') or '',
        'Source':    l.get('source') or '',
        'Statut':    l['status'],
        'Date':      l['created_at'],
        'Lien WA':   f"https://wa.me/{l['whatsapp'].replace(' ','').replace('+','')}",
    } for l in leads]
    df = pd.DataFrame(data) if data else pd.DataFrame()
    output = BytesIO()
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    if fmt == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Leads Trading')
            try:
                from openpyxl.styles import PatternFill, Font, Alignment
                ws = writer.sheets['Leads Trading']
                fill = PatternFill("solid", fgColor="0D3B2D")
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = Font(bold=True, color="00C77A")
                    cell.alignment = Alignment(horizontal='center')
                for col in ws.columns:
                    w = max(len(str(c.value or '')) for c in col) + 4
                    ws.column_dimensions[col[0].column_letter].width = min(w, 40)
            except Exception:
                pass
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name=f'leads_trading_{ts}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name=f'leads_trading_{ts}.csv',
                         mimetype='text/csv')


@app.route('/api/leads/stats')
@login_required
@admin_required
def leads_stats_api():
    pipeline_status = [{"$group": {"_id": "$status", "cnt": {"$sum": 1}}}]
    pipeline_level  = [{"$group": {"_id": "$level_selected", "cnt": {"$sum": 1}}}]
    rows       = {r['_id']: r['cnt'] for r in get_col('training_leads').aggregate(pipeline_status)}
    by_level   = {r['_id']: r['cnt'] for r in get_col('training_leads').aggregate(pipeline_level)}
    total = sum(rows.values())
    paye  = rows.get('Payé', 0)
    return jsonify({
        'total':           total,
        'par_statut':      rows,
        'par_formation':   by_level,
        'taux_conversion': round(paye / total * 100, 1) if total else 0,
    })

# ══════════════════════════════════════════════════════════════
# AGENDA — Email Gmail + Scheduler + Routes
# ══════════════════════════════════════════════════════════════

def _build_agenda_email_html(event: dict, minutes_before: int) -> str:
    etype  = event.get('event_type', 'personnel')
    cfg    = AGENDA_EVENT_COLORS.get(etype, AGENDA_EVENT_COLORS['personnel'])
    color  = cfg['bg']
    icon   = cfg['icon']
    label  = cfg['label']
    title  = event.get('title', '(Sans titre)')
    desc   = event.get('description') or ''
    loc    = event.get('location') or ''
    notes  = event.get('notes') or ''
    start  = (event.get('start_datetime') or '')[:16].replace('T', ' à ')
    end    = (event.get('end_datetime')   or '')[:16].replace('T', ' à ')
    if minutes_before >= 1440:
        remind_txt = f"{minutes_before // 1440} jour(s)"
    elif minutes_before >= 60:
        remind_txt = f"{minutes_before // 60} heure(s)"
    else:
        remind_txt = f"{minutes_before} minute(s)"
    loc_row   = f'<tr><td style="padding:8px 0;color:#888;width:100px;">📍 Lieu</td><td style="padding:8px 0;color:#e0e0e0;">{loc}</td></tr>' if loc else ''
    notes_blk = f'<div style="background:#1a2a3a;border-radius:8px;padding:16px;margin-top:18px;"><p style="color:#888;font-size:11px;margin:0 0 6px;">📝 Notes</p><p style="color:#ccc;font-size:13px;margin:0;">{notes}</p></div>' if notes else ''
    desc_blk  = f'<p style="color:#aaa;margin:4px 0 0;font-size:14px;">{desc}</p>' if desc else ''
    return f'''<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8"/></head>
<body style="margin:0;padding:0;background:#0a0f1a;font-family:'Segoe UI',Arial,sans-serif;">
<div style="max-width:580px;margin:0 auto;padding:24px;">
  <div style="background:linear-gradient(135deg,#0d1b2a,#1a2a3a);border-radius:16px 16px 0 0;padding:32px 28px;text-align:center;border-bottom:3px solid {color};">
    <div style="font-size:42px;margin-bottom:10px;">⏰</div>
    <h1 style="color:#fff;margin:0;font-size:22px;font-weight:800;">Rappel d'Agenda</h1>
    <p style="color:{color};margin:8px 0 0;font-size:14px;font-weight:600;">{icon} {label} · dans {remind_txt}</p>
  </div>
  <div style="background:#111827;padding:28px;border-radius:0 0 16px 16px;border:1px solid #1e2a3a;border-top:none;">
    <div style="background:linear-gradient(135deg,{color}18,{color}08);border:1px solid {color}44;border-left:5px solid {color};border-radius:10px;padding:20px;margin-bottom:22px;">
      <h2 style="color:#fff;margin:0 0 4px;font-size:20px;font-weight:700;">{title}</h2>
      {desc_blk}
    </div>
    <table style="width:100%;border-collapse:collapse;margin-bottom:4px;">
      <tr><td style="padding:8px 0;color:#888;width:100px;">🕐 Début</td><td style="padding:8px 0;color:#e0e0e0;font-weight:600;">{start}</td></tr>
      <tr><td style="padding:8px 0;color:#888;">🏁 Fin</td><td style="padding:8px 0;color:#e0e0e0;">{end}</td></tr>
      {loc_row}
    </table>
    {notes_blk}
    <div style="border-top:1px solid #1e2a3a;padding-top:14px;text-align:center;">
      <p style="color:#333;font-size:11px;margin:0;">Kengni Finance · Rappel automatique</p>
    </div>
  </div>
</div>
</body></html>'''


def _send_agenda_email(event: dict, minutes_before: int) -> bool:
    cfg = GMAIL_CONFIG
    try:
        msg = MIMEMultipart('alternative')
        h = f"{'%dh' % (minutes_before//60) if minutes_before >= 60 else '%dmin' % minutes_before}"
        msg['Subject'] = f"⏰ Rappel dans {h} : {event['title']}"
        msg['From']    = f"{cfg['sender_name']} <{cfg['sender_email']}>"
        msg['To']      = cfg['receiver_email']
        text = (f"RAPPEL — {event['title']}\n"
                f"Début : {(event.get('start_datetime') or '')[:16]}\n"
                f"Fin   : {(event.get('end_datetime') or '')[:16]}\n"
                f"Lieu  : {event.get('location') or 'Non précisé'}\n\n"
                f"{event.get('description') or ''}\n\n---\nKengni Finance")
        msg.attach(MIMEText(text, 'plain', 'utf-8'))
        msg.attach(MIMEText(_build_agenda_email_html(event, minutes_before), 'html', 'utf-8'))
        with smtplib.SMTP(cfg['smtp_host'], cfg['smtp_port']) as s:
            s.ehlo(); s.starttls(); s.login(cfg['sender_email'], cfg['smtp_password'])
            s.sendmail(cfg['sender_email'], cfg['receiver_email'], msg.as_string())
        print(f"[Agenda] ✅ Email envoyé : {event['title']} ({minutes_before}min avant)")
        return True
    except smtplib.SMTPAuthenticationError:
        print(f"[Agenda] ❌ Auth Gmail échouée")
        return False
    except Exception as e:
        print(f"[Agenda] ❌ Erreur email : {e}")
        return False


def _agenda_check_reminders():
    """Vérifie et envoie les rappels toutes les 60 secondes.
    NOTE: Sur Vercel (serverless), ce thread ne s'exécute pas.
    Pour Vercel, utilisez MongoDB Atlas Scheduled Triggers ou un cron job externe.
    """
    while True:
        try:
            now = datetime.now()
            window = (now + timedelta(hours=48)).isoformat()
            events = docs_to_list(get_col('agenda_events').find({
                "status": "active",
                "start_datetime": {"$gte": now.isoformat(), "$lte": window},
                "$or": [{"email_reminder": 1}, {"app_reminder": 1}]
            }))
            for ev in events:
                try:
                    start_dt  = datetime.fromisoformat(ev['start_datetime'])
                    remind_at = start_dt - timedelta(minutes=ev['reminder_minutes'])
                    diff = abs((now - remind_at).total_seconds())
                    if diff > 90:
                        continue
                    # Anti-doublon
                    cutoff = (now - timedelta(minutes=3)).isoformat()
                    already_sent = get_col('agenda_reminders_sent').find_one({
                        "event_id": ev['id'],
                        "sent_at": {"$gte": cutoff}
                    })
                    if already_sent:
                        continue
                    if ev.get('email_reminder'):
                        ok = _send_agenda_email(ev, ev['reminder_minutes'])
                        if ok:
                            rid = get_next_id("agenda_reminders_sent")
                            get_col('agenda_reminders_sent').insert_one({
                                "id": rid, "event_id": ev['id'],
                                "sent_at": now.isoformat(), "method": "email"
                            })
                    if ev.get('app_reminder'):
                        h     = ev['reminder_minutes']
                        label = f"{h//60}h" if h >= 60 else f"{h}min"
                        nid   = get_next_id("notifications")
                        get_col('notifications').insert_one({
                            "id": nid,
                            "user_id": ev['user_id'],
                            "type": 'warning' if h <= 15 else 'info',
                            "title": f"⏰ Rappel dans {label} : {ev['title']}",
                            "message": f"Votre événement commence à {ev['start_datetime'][11:16]}.",
                            "action_url": '/agenda',
                            "is_read": 0,
                            "created_at": now.isoformat()
                        })
                        rid = get_next_id("agenda_reminders_sent")
                        get_col('agenda_reminders_sent').insert_one({
                            "id": rid, "event_id": ev['id'],
                            "sent_at": now.isoformat(), "method": "app"
                        })
                except Exception as ex:
                    print(f"[Agenda] Erreur traitement event #{ev.get('id')}: {ex}")
        except Exception as e:
            print(f"[Agenda Scheduler] Erreur : {e}")
        time.sleep(60)


def start_agenda_scheduler():
    t = threading.Thread(target=_agenda_check_reminders, daemon=True, name='AgendaScheduler')
    t.start()
    print("✅ Agenda Scheduler démarré (Gmail → iCloud, toutes les 60s)")


@app.route('/agenda')
@login_required
def agenda():
    user_id = session['user_id']
    now     = datetime.now()
    start_w = (now - timedelta(days=60)).strftime('%Y-%m-%d')
    end_w   = (now + timedelta(days=90)).strftime('%Y-%m-%d')
    events = docs_to_list(get_col('agenda_events').find({
        "user_id": user_id,
        "status": {"$ne": "cancelled"},
        "start_datetime": {"$gte": start_w, "$lte": end_w}
    }, sort=[("start_datetime", 1)]))
    next_event_doc = doc_to_dict(get_col('agenda_events').find_one({
        "user_id": user_id,
        "status": "active",
        "start_datetime": {"$gte": now.isoformat()}
    }, sort=[("start_datetime", 1)]))
    stats_pipeline = [
        {"$match": {"user_id": user_id, "status": "active",
                    "start_datetime": {"$gte": now.strftime('%Y-%m-%d')}}},
        {"$group": {"_id": "$event_type", "cnt": {"$sum": 1}}}
    ]
    stats_raw = {r['_id']: r['cnt'] for r in get_col('agenda_events').aggregate(stats_pipeline)}
    fc_events = []
    for e in events:
        cfg = AGENDA_EVENT_COLORS.get(e['event_type'], AGENDA_EVENT_COLORS['personnel'])
        fc_events.append({
            'id': e['id'], 'title': e['title'],
            'start': e['start_datetime'], 'end': e['end_datetime'],
            'allDay': bool(e['all_day']),
            'backgroundColor': cfg['bg'], 'borderColor': cfg['border'],
            'extendedProps': {
                'type': e['event_type'], 'icon': cfg['icon'],
                'description': e['description'] or '',
                'location': e['location'] or '',
                'notes': e['notes'] or '',
                'reminder': e['reminder_minutes'],
            }
        })
    return render_template('agenda.html',
        events_json  = json.dumps(fc_events),
        events       = events,
        next_event   = next_event_doc,
        stats        = stats_raw,
        event_colors = AGENDA_EVENT_COLORS,
        now          = now,
    )


@app.route('/api/agenda/events', methods=['POST'])
@login_required
def agenda_create_event():
    user_id = session['user_id']
    data    = request.get_json() or {}
    title      = (data.get('title') or '').strip()
    event_type = data.get('event_type', 'personnel')
    start_dt   = data.get('start_datetime', '')
    end_dt     = data.get('end_datetime', '')
    if not title or not start_dt or not end_dt:
        return jsonify({'success': False, 'error': 'Champs requis manquants'}), 400
    if event_type not in AGENDA_EVENT_COLORS:
        event_type = 'personnel'
    eid = get_next_id("agenda_events")
    get_col('agenda_events').insert_one({
        "id": eid,
        "user_id": user_id,
        "title": title,
        "description": data.get('description', ''),
        "event_type": event_type,
        "start_datetime": start_dt,
        "end_datetime": end_dt,
        "all_day": int(data.get('all_day', 0)),
        "recurrence": data.get('recurrence', 'none'),
        "reminder_minutes": int(data.get('reminder_minutes', 30)),
        "email_reminder": int(data.get('email_reminder', 1)),
        "app_reminder": int(data.get('app_reminder', 1)),
        "location": data.get('location', ''),
        "notes": data.get('notes', ''),
        "status": "active",
        "created_at": _now_iso(),
        "updated_at": _now_iso()
    })
    nid = get_next_id("notifications")
    get_col('notifications').insert_one({
        "id": nid,
        "user_id": user_id,
        "type": "success",
        "title": f"📅 Événement créé : {title}",
        "message": f'Planifié le {start_dt[:16].replace("T"," à ")}. Rappel dans {data.get("reminder_minutes",30)} min.',
        "action_url": '/agenda',
        "is_read": 0,
        "created_at": _now_iso()
    })
    return jsonify({'success': True, 'event_id': eid})


@app.route('/api/agenda/events')
@login_required
def agenda_get_events():
    user_id = session['user_id']
    start   = request.args.get('start', (datetime.now() - timedelta(days=30)).isoformat())
    end     = request.args.get('end',   (datetime.now() + timedelta(days=90)).isoformat())
    events  = docs_to_list(get_col('agenda_events').find({
        "user_id": user_id,
        "status": {"$ne": "cancelled"},
        "start_datetime": {"$gte": start[:10], "$lte": end[:10]}
    }, sort=[("start_datetime", 1)]))
    result = []
    for e in events:
        cfg = AGENDA_EVENT_COLORS.get(e['event_type'], AGENDA_EVENT_COLORS['personnel'])
        result.append({
            'id': e['id'], 'title': e['title'],
            'start': e['start_datetime'], 'end': e['end_datetime'],
            'allDay': bool(e['all_day']),
            'backgroundColor': cfg['bg'], 'borderColor': cfg['border'],
            'extendedProps': {
                'type': e['event_type'], 'icon': cfg['icon'],
                'description': e['description'] or '',
                'location': e['location'] or '',
                'reminder': e['reminder_minutes']
            }
        })
    return jsonify(result)


@app.route('/api/agenda/events/<event_id>', methods=['PUT', 'PATCH'])
@login_required
def agenda_update_event(event_id):
    user_id = session['user_id']
    data    = request.get_json() or {}
    existing = get_col('agenda_events').find_one({"id": int(event_id), "user_id": user_id})
    if not existing:
        return jsonify({'success': False, 'error': 'Non trouvé'}), 404
    fields = ['title','description','event_type','start_datetime','end_datetime',
              'all_day','recurrence','reminder_minutes','email_reminder',
              'app_reminder','location','notes','status']
    updates = {"updated_at": _now_iso()}
    for f in fields:
        if f in data:
            updates[f] = data[f]
    if len(updates) > 1:
        get_col('agenda_events').update_one(
            {"id": int(event_id), "user_id": user_id},
            {"$set": updates}
        )
    return jsonify({'success': True})


@app.route('/api/agenda/events/<event_id>', methods=['DELETE'])
@login_required
def agenda_delete_event(event_id):
    user_id = session['user_id']
    get_col('agenda_events').update_one(
        {"id": int(event_id), "user_id": user_id},
        {"$set": {"status": "cancelled"}}
    )
    return jsonify({'success': True})


@app.route('/api/agenda/today')
@login_required
def agenda_today():
    user_id = session['user_id']
    today   = datetime.now().strftime('%Y-%m-%d')
    events  = docs_to_list(get_col('agenda_events').find({
        "user_id": user_id,
        "status": "active",
        "start_datetime": {"$regex": f"^{today}"}
    }, sort=[("start_datetime", 1)]))
    return jsonify({'events': events, 'count': len(events), 'date': today})


@app.route('/api/agenda/test-email', methods=['POST'])
@login_required
def agenda_test_email():
    now = datetime.now()
    fake_event = {
        'id': 0,
        'user_id': session['user_id'],
        'title': '✅ Test de configuration — Kengni Finance Agenda',
        'event_type': 'trading',
        'description': 'Ceci est un email de test envoyé depuis Kengni Finance pour vérifier que la configuration Gmail fonctionne.',
        'start_datetime': (now + timedelta(minutes=30)).strftime('%Y-%m-%dT%H:%M:%S'),
        'end_datetime':   (now + timedelta(minutes=90)).strftime('%Y-%m-%dT%H:%M:%S'),
        'location':       'Kengni Finance',
        'notes':          f'Test effectué le {now.strftime("%d/%m/%Y à %H:%M")} par {session.get("username","admin")}.',
    }
    ok = _send_agenda_email(fake_event, 30)
    if ok:
        return jsonify({'success': True,
                        'message': f'✅ Email envoyé avec succès à {GMAIL_CONFIG["receiver_email"]} !',
                        'from': GMAIL_CONFIG['sender_email'],
                        'to':   GMAIL_CONFIG['receiver_email']})
    return jsonify({'success': False,
                    'message': '❌ Échec — le mot de passe d\'application Gmail est incorrect ou manquant.',
                    'help': 'Générez un mot de passe sur https://myaccount.google.com/apppasswords'}), 500


# ══════════════════════════════════════════════════════════════
# DÉMARRAGE
# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    init_db()
    start_agenda_scheduler()
    print("=" * 60)
    print("🚀 Kengni Finance v2.0 - Démarrage (MongoDB)")
    print("=" * 60)
    print("📊 Application de gestion financière et trading avec IA")
    print("🌐 URL: http://localhost:5001")
    print("👤 Email:", os.environ.get('ADMIN_EMAIL', 'fabrice.kengni@icloud.com'))
    print("=" * 60)
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))